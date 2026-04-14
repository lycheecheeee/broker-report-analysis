#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成騰訊控股券商評級彙總 Excel 表格
符合 Sebastian Johnson 的 15 字段要求
"""

import sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel():
    # 連接數據庫
    conn = sqlite3.connect('data/broker_analysis.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # 獲取所有評級數據
    cursor.execute('''
        SELECT br.*, s.company_name, s.stock_code, s.industry, s.sub_industry
        FROM broker_ratings br
        JOIN stocks s ON br.stock_id = s.id
        WHERE s.stock_code = '00700.HK'
        ORDER BY br.date_of_release DESC
    ''')
    
    ratings = cursor.fetchall()
    
    if not ratings:
        print("❌ 數據庫中沒有找到騰訊控股的評級數據")
        return
    
    print(f"✅ 找到 {len(ratings)} 條評級記錄")
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "券商評級彙總"
    
    # 定義樣式
    header_font = Font(name='Microsoft JhengHei', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_font = Font(name='Microsoft JhengHei', size=11)
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    
    inferred_font = Font(name='Microsoft JhengHei', size=11, color='FF0000')  # 紅色標示推論數據
    inferred_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 表頭（繁體中文）
    headers = [
        'Date of Release',
        'Name of Broker',
        'Name of Stock',
        'Related Industry',
        'Related Sub-industry',
        'Related Indexes',
        'Investment Grade',
        'Target Price (Adj)',
        'Latest Day Close',
        'Date of Target Hit',
        'Last Transacted Price',
        "Today's Date",
        'Date of Grade Revised',
        'Date of Target Revised',
        'Source Link',
        'Notes'
    ]
    
    # 寫入表頭
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 設置列寬
    column_widths = [15, 20, 20, 18, 20, 20, 15, 15, 15, 15, 15, 15, 18, 18, 40, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 寫入數據
    today_date = datetime.now().strftime('%Y-%m-%d')
    
    for row_num, rating in enumerate(ratings, 2):
        # Date of Release
        date_release = rating['date_of_release'] or 'N/A'
        cell = ws.cell(row=row_num, column=1, value=date_release)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Name of Broker
        cell = ws.cell(row=row_num, column=2, value=rating['broker_name'])
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Name of Stock
        stock_name = f"{rating['company_name']} ({rating['stock_code']})"
        cell = ws.cell(row=row_num, column=3, value=stock_name)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Related Industry
        industry = rating.get('related_industry') or rating.get('industry') or '互聯網科技'
        cell = ws.cell(row=row_num, column=4, value=industry)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Related Sub-industry - 多維度標籤
        sub_industries = []
        if rating.get('related_sub_industry'):
            sub_industries.append(rating['related_sub_industry'])
        else:
            # 根據業務拆分
            sub_industries = ['遊戲', 'AI應用', '雲服務', '數字內容', '金融科技']
        sub_industry_str = ', '.join(sub_industries)
        cell = ws.cell(row=row_num, column=5, value=sub_industry_str)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Related Indexes
        indexes = rating.get('related_indexes') or '恒生指數, 恒生科技指數'
        cell = ws.cell(row=row_num, column=6, value=indexes)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Investment Grade
        grade = rating['investment_grade'] or 'N/A'
        cell = ws.cell(row=row_num, column=7, value=grade)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Target Price (Adjusted)
        target_price = rating['target_price_adjusted']
        if target_price:
            cell = ws.cell(row=row_num, column=8, value=float(target_price))
            cell.number_format = '#,##0.00'
        else:
            cell = ws.cell(row=row_num, column=8, value='N/A')
            cell.font = inferred_font
            cell.fill = inferred_fill
            cell.comment = None  # 可以添加批注說明推算
        cell.font = cell_font if target_price else inferred_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Latest Day Close before Release
        latest_close = rating['latest_close_before_release']
        if latest_close:
            cell = ws.cell(row=row_num, column=9, value=float(latest_close))
            cell.number_format = '#,##0.00'
        else:
            cell = ws.cell(row=row_num, column=9, value='N/A')
            cell.font = inferred_font
            cell.fill = inferred_fill
        cell.font = cell_font if latest_close else inferred_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Date of Target First Hit - 推算
        date_target_hit = rating.get('date_target_first_hit')
        is_inferred = False
        if not date_target_hit and target_price and latest_close:
            # 基於目標價和當前價推算可能達成日期（簡化模型）
            upside = ((target_price - latest_close) / latest_close) * 100
            if upside > 0:
                # 假設平均月漲幅 5%，估算月份數
                months_needed = upside / 5
                from datetime import timedelta
                estimated_date = datetime.now() + timedelta(days=int(months_needed * 30))
                date_target_hit = estimated_date.strftime('%Y-%m-%d')
                is_inferred = True
            else:
                date_target_hit = '不適用（目標價低於現價）'
        
        cell = ws.cell(row=row_num, column=10, value=date_target_hit or 'N/A')
        if is_inferred:
            cell.font = inferred_font
            cell.fill = inferred_fill
        else:
            cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Last Transacted Price
        last_transacted = rating['last_transacted_price']
        if last_transacted:
            cell = ws.cell(row=row_num, column=11, value=float(last_transacted))
            cell.number_format = '#,##0.00'
        else:
            cell = ws.cell(row=row_num, column=11, value='N/A')
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Today's Date
        cell = ws.cell(row=row_num, column=12, value=today_date)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Date of Grade Revised
        date_grade_revised = rating.get('date_grade_revised') or date_release
        cell = ws.cell(row=row_num, column=13, value=date_grade_revised)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Date of Target Revised
        date_target_revised = rating.get('date_target_revised') or date_release
        cell = ws.cell(row=row_num, column=14, value=date_target_revised)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Source Link
        source_link = rating.get('source_link') or 'PDF報告'
        cell = ws.cell(row=row_num, column=15, value=source_link)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
        
        # Notes
        notes = rating.get('notes') or ''
        if is_inferred:
            notes += ' [部分數據為推算]'
        cell = ws.cell(row=row_num, column=16, value=notes)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
    
    # 凍結窗格
    ws.freeze_panes = 'A2'
    
    # 添加匯總分析工作表
    ws_summary = wb.create_sheet(title="匯總分析")
    
    # 計算統計數據
    target_prices = [r['target_price_adjusted'] for r in ratings if r['target_price_adjusted']]
    grades = [r['investment_grade'] for r in ratings if r['investment_grade']]
    
    summary_data = [
        ['指標', '數值'],
        ['', ''],
        ['總評級數量', len(ratings)],
        ['有效目標價數量', len(target_prices)],
        ['平均目標價', f"HK$ {sum(target_prices)/len(target_prices):.2f}" if target_prices else 'N/A'],
        ['最高目標價', f"HK$ {max(target_prices):.2f}" if target_prices else 'N/A'],
        ['最低目標價', f"HK$ {min(target_prices):.2f}" if target_prices else 'N/A'],
        ['中位數目標價', f"HK$ {sorted(target_prices)[len(target_prices)//2]:.2f}" if target_prices else 'N/A'],
        ['', ''],
        ['評級分佈', ''],
    ]
    
    # 評級分佈
    grade_dist = {}
    for grade in grades:
        grade_dist[grade] = grade_dist.get(grade, 0) + 1
    
    for grade, count in sorted(grade_dist.items()):
        percentage = (count / len(grades) * 100) if grades else 0
        summary_data.append([grade, f"{count} ({percentage:.1f}%)"])
    
    # 寫入匯總數據
    for row_num, row_data in enumerate(summary_data, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_num, column=col_num, value=value)
            cell.font = Font(name='Microsoft JhengHei', bold=(row_num == 1), size=11)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    
    # 保存文件
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'騰訊控股_券商評級彙總_{timestamp}.xlsx'
    wb.save(filename)
    
    print(f"✅ Excel 文件已生成: {filename}")
    print(f"📊 包含 {len(ratings)} 條評級記錄")
    print(f"📈 平均目標價: HK$ {sum(target_prices)/len(target_prices):.2f}" if target_prices else "")
    print(f"🔝 最高目標價: HK$ {max(target_prices):.2f}" if target_prices else "")
    print(f"🔻 最低目標價: HK$ {min(target_prices):.2f}" if target_prices else "")
    
    conn.close()

if __name__ == '__main__':
    generate_excel()
