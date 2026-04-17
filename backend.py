from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import hashlib
import json
import os
import logging
from datetime import datetime

# 配置日誌系統 - 生產環境只記錄 WARNING 及以上
logger = logging.getLogger(__name__)
if os.environ.get('FLASK_ENV') == 'development':
    logging.basicConfig(level=logging.DEBUG, format='%(asctime)s [%(levelname)s] %(message)s')
else:
    logging.basicConfig(level=logging.WARNING, format='%(asctime)s [%(levelname)s] %(message)s')

# 數據庫支持 - SQLite 或 PostgreSQL
try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
    USE_POSTGRESQL = True
    logger.info("PostgreSQL support available")
except ImportError:
    USE_POSTGRESQL = False
    import sqlite3
    logger.warning("Using SQLite (PostgreSQL not available)")

# 可選導入 - 避免 Vercel 環境崩潰
try:
    import jwt
    JWT_AVAILABLE = True
except ImportError:
    JWT_AVAILABLE = False
    logger.warning("PyJWT not available")

try:
    import PyPDF2
    PYPDF2_AVAILABLE = True
except ImportError:
    PYPDF2_AVAILABLE = False
    logger.warning("PyPDF2 not available")

import re
import requests

app = Flask(__name__, static_folder='.', static_url_path='')

# 簡化 CORS 配置 - 避免 Vercel 環境問題
try:
    CORS(app)
    logger.info("CORS initialized")
except Exception as e:
    logger.error(f"CORS initialization error: {e}")

app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 最大上傳 50MB

# 添加請求日誌中間件
@app.before_request
def log_request():
    """請求日誌 - 僅開發環境啟用"""
    if os.environ.get('FLASK_ENV') == 'development':
        logger.debug(f"{request.method} {request.path}")
        if request.form:
            logger.debug(f"Form data: {dict(request.form)}")


# 配置
import secrets
SECRET_KEY = os.environ.get('SECRET_KEY', '').strip()
if not SECRET_KEY:
    SECRET_KEY = secrets.token_hex(32)
    logger.warning("SECRET_KEY not set, generated a random one. Set it in Vercel env for production.")

# Supabase 配置 - 清理可能的換行符和空白
SUPABASE_URL = os.environ.get('SUPABASE_URL', '').strip()
SUPABASE_KEY = os.environ.get('SUPABASE_KEY', '').strip()

# 強制要求 Supabase 配置 - 公開工具必須使用雲端數據庫
if not SUPABASE_URL or not SUPABASE_KEY:
    raise EnvironmentError(
        "Missing required environment variables: SUPABASE_URL and SUPABASE_KEY must be set.\n"
        "This is a public tool that requires cloud database persistence.\n"
        "Please configure these in Vercel Dashboard > Settings > Environment Variables"
    )

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), '700')

# 所有 API Key 讀取時統一 strip，防止隱藏空白/換行符導致 HTTP header 無效
OPENROUTER_API_KEY = os.environ.get('OPENROUTER_API_KEY', '').strip()
OPENROUTER_API_URL = 'https://openrouter.ai/api/v1/chat/completions'

# NVIDIA NIM API 配置（優先使用）
NVIDIA_API_KEY = os.environ.get('NVIDIA_API_KEY', '').strip()
NVIDIA_API_URL = 'https://integrate.api.nvidia.com/v1/chat/completions'
NVIDIA_MODEL = 'meta/llama-3.1-8b-instruct'  # 使用 Llama 3.1 8B（速度快，免費）

# 數據庫模式：強制使用 Supabase
DB_MODE = 'supabase'

# 確保上傳文件夾存在
try:
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)
    logger.info(f"Upload folder ready: {UPLOAD_FOLDER}")
except Exception as e:
    logger.error(f"Upload folder creation error: {e}")

# SQLite 數據庫路徑（僅本地開發使用）
DATABASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'broker_reports.db')

# 數據庫初始化標誌
_db_initialized = False

def supabase_request(method, table, data=None, query_params=None):
    """Supabase REST API 請求輔助函數"""
    if DB_MODE != 'supabase' or not SUPABASE_KEY:
        logger.error(f"Supabase not configured: DB_MODE={DB_MODE}, KEY_SET={bool(SUPABASE_KEY)}")
        return None
    
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    headers = {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'return=representation'  # 返回插入/更新後的數據
    }
    
    try:
        logger.debug(f"Supabase {method} request to {url}")
        if method == 'GET':
            if query_params:
                url += f"?{query_params}"
            response = requests.get(url, headers=headers, timeout=10)
        elif method == 'POST':
            logger.debug(f"POST data keys: {list(data.keys()) if data else 'None'}")
            response = requests.post(url, headers=headers, json=data, timeout=10)
        elif method == 'PATCH':
            if query_params:
                url += f"?{query_params}"
            response = requests.patch(url, headers=headers, json=data, timeout=10)
        elif method == 'DELETE':
            if query_params:
                url += f"?{query_params}"
            response = requests.delete(url, headers=headers, timeout=10)
        else:
            return None
        
        logger.debug(f"Supabase response status: {response.status_code}")
        if response.status_code in [200, 201, 204]:
            if response.status_code == 204:
                return []
            result = response.json()
            logger.debug(f"Supabase response data: {len(result) if isinstance(result, list) else 'object'} items")
            return result
        else:
            logger.error(f"Supabase error {response.status_code}: {response.text[:500]}")
            return None
    except Exception as e:
        logger.error(f"Supabase request failed: {type(e).__name__}: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return None

def init_db():
    """初始化數據庫"""
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    
    # 用戶表
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        email TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # PDF分析結果表
    c.execute('''CREATE TABLE IF NOT EXISTS analysis_results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        company_name TEXT,
        stock_code TEXT,
        pdf_filename TEXT NOT NULL,
        broker_name TEXT,
        rating TEXT,
        target_price REAL,
        current_price REAL,
        upside_potential REAL,
        ai_summary TEXT,
        key_points TEXT,
        risks TEXT,
        prompt_used TEXT,
        chart_path TEXT,
        audio_path TEXT,
        is_public BOOLEAN DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (user_id) REFERENCES users(id)
    )''')
    
    # Migration: 添加缺失字段(如果已存在則忽略)
    migration_columns = [
        ('company_name', 'TEXT'),
        ('stock_code', 'TEXT'),
        ('key_points', 'TEXT'),
        ('risks', 'TEXT'),
        ('chart_path', 'TEXT'),
        ('audio_path', 'TEXT'),
        ('is_public', 'BOOLEAN DEFAULT 0')
    ]
    
    for col_name, col_type in migration_columns:
        try:
            c.execute(f"ALTER TABLE analysis_results ADD COLUMN {col_name} {col_type}")
            logger.info(f"Added column: {col_name}")
        except:
            pass  # 字段已存在,忽略
    
    # 用戶反饋表
    c.execute('''CREATE TABLE IF NOT EXISTS feedback (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        analysis_id INTEGER,
        rating INTEGER CHECK(rating >= 1 AND rating <= 5),
        comment TEXT,
        suggestions TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (user_id) REFERENCES users(id),
        FOREIGN KEY (analysis_id) REFERENCES analysis_results(id)
    )''')
    
    # 自定義Prompt模板表
    c.execute('''CREATE TABLE IF NOT EXISTS prompt_templates (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        template_name TEXT NOT NULL,
        prompt_text TEXT NOT NULL,
        is_default BOOLEAN DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (user_id) REFERENCES users(id)
    )''')
    
    # 股票表
    c.execute('''CREATE TABLE IF NOT EXISTS stocks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_name TEXT NOT NULL,
        stock_code TEXT UNIQUE NOT NULL,
        industry TEXT,
        sub_industry TEXT,
        current_price REAL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # 券商評級表(完整 15 個字段)
    c.execute('''CREATE TABLE IF NOT EXISTS broker_ratings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        stock_id INTEGER NOT NULL,
        
        -- 基本信息
        date_of_release TEXT,                    -- 1. Date of Release
        broker_name TEXT NOT NULL,               -- 2. Name of Broker
        stock_name TEXT,                         -- 3. Name of Stock
        related_industry TEXT,                   -- 4. Related Industry
        related_sub_industry TEXT,               -- 5. Related Sub-industry
        related_indexes TEXT,                    -- 6. Related Indexes
        
        -- 投資評級與目標價
        investment_grade TEXT,                   -- 7. Investment Grade (Buy/Hold/Sell)
        target_price_adjusted REAL,              -- 8. Target Price (Adjusted)
        investment_horizon TEXT DEFAULT '12 months',  -- 9. Investment Horizon (default 12 months)
        
        -- 價格信息
        latest_close_before_release REAL,        -- 10. Latest Day Close before Release (Adjusted)
        date_target_first_hit TEXT,              -- 11. Date of Target First Hit
        last_transacted_price REAL,              -- 12. Last Transacted/Closing Price as of Today
        today_date TEXT,                         -- 13. Today's Date
        
        -- 修訂日期
        date_grade_revised TEXT,                 -- 14. Date of Investment Grade Revised/Extended
        date_target_revised TEXT,                -- 15. Date of Target Price Revised/Extended
        
        -- 其他
        source_link TEXT,
        notes TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (stock_id) REFERENCES stocks(id)
    )''')
    
    conn.commit()
    conn.close()

def ensure_db_initialized():
    """確保數據庫已初始化（懶加載）"""
    global _db_initialized
    if not _db_initialized:
        try:
            if DB_MODE == 'supabase':
                # Supabase 模式：測試連接
                result = supabase_request('GET', 'analysis_results', query_params='limit=1')
                logger.info(f"Supabase connected (mode: {DB_MODE})")
            else:
                # SQLite/內存模式
                init_db()
                logger.info(f"Database initialized (mode: {DB_MODE})")
            _db_initialized = True
        except Exception as e:
            logger.error(f"Database initialization error: {e}")
            # 即使失敗也標記為已初始化，避免重複嘗試
            _db_initialized = True

# Auth 系統已移除 - 公開工具無需登入
# 所有 API 使用固定 user_id = 1

def get_active_records(all_results):
    """從全部 record 中只取每個 pdf_filename 最新的一條（active record）。
    舊 record 自動成為 archive，唔刪除。
    
    Args:
        all_results: Supabase 返回嘅全部 record list
    Returns:
        list: 每個 pdf_filename 最新的一條 record
    """
    if not all_results:
        return []
    
    from collections import defaultdict
    groups = defaultdict(list)
    for r in all_results:
        fname = r.get('pdf_filename', '')
        groups[fname].append(r)
    
    # 每個 group 只保留 created_at 最新的一條
    active = []
    for fname, records in groups.items():
        latest = max(records, key=lambda x: x.get('created_at', ''))
        active.append(latest)
    
    return active

def get_archived_records(all_results):
    """從全部 record 中取所有非最新嘅 record（archived record）。
    
    Args:
        all_results: Supabase 返回嘅全部 record list
    Returns:
        list: 所有非最新嘅 record，按 created_at 降序排列
    """
    if not all_results:
        return []
    
    from collections import defaultdict
    groups = defaultdict(list)
    for r in all_results:
        fname = r.get('pdf_filename', '')
        groups[fname].append(r)
    
    # 每個 group 取最新一條嘅 id
    latest_ids = set()
    for fname, records in groups.items():
        latest = max(records, key=lambda x: x.get('created_at', ''))
        latest_ids.add(latest.get('id'))
    
    # 所有唔係最新嘅 record
    archived = [r for r in all_results if r.get('id') not in latest_ids]
    archived.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    
    return archived

def parse_pdf(pdf_path):
    """解析PDF文件"""
    try:
        logger.debug(f"Parsing PDF: {pdf_path}")
        file_size = os.path.getsize(pdf_path)
        
        # 檢查文件大小，如果太小則跳過
        if file_size < 1000:
            logger.warning(f"File too small ({file_size} bytes), may be corrupted")
            return None
        
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file, strict=False)  # 設置strict=False以容忍一些錯誤
            logger.debug(f"PDF pages: {len(pdf_reader.pages)}")
            
            text = ''
            for i, page in enumerate(pdf_reader.pages[:2]):  # 只讀前2頁（優化性能）
                try:
                    page_text = page.extract_text()
                    if page_text:
                        logger.debug(f"Page {i+1} extracted, length: {len(page_text)}")
                        text += page_text + '\n'
                    else:
                        logger.debug(f"Page {i+1} has no text content")
                except Exception as page_error:
                    logger.warning(f"Page {i+1} extraction failed: {page_error}")
                    continue
            
            logger.debug(f"Total text length: {len(text)}")
            return text if text.strip() else None
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        logger.error(f"PDF parse error: {e}\n{error_detail}")
        return None

def extract_broker_info(text):
    """從文本提取券商信息"""
    broker_mapping = {
        'BOA': '美國銀行',
        'Bank of America': '美國銀行',
        'CICC': '中金公司',
        'Citigroup': '花旗',
        'CLSA': '里昂證券',
        'CMB': '招銀國際',
        'CMS': '招商證券',
        'Daiwa': '大和資本',
        'Deutsche Bank': '德意志銀行',
        'JPMorgan': '摩根大通',
        'Macquarie': '麥格理',
        'Morgan Stanley': '摩根士丹利',
        'Nomura': '野村證券',
        'UBS': '瑞銀'
    }
    
    broker_name = '未知券商'
    for key, value in broker_mapping.items():
        if key.lower() in text.lower():
            broker_name = value
            break
    
    # 提取評級
    rating = '未明確'
    if any(kw in text.lower() for kw in ['buy', '買入']):
        rating = '買入'
    elif any(kw in text.lower() for kw in ['overweight', '增持']):
        rating = '增持'
    elif any(kw in text.lower() for kw in ['outperform', '跑贏']):
        rating = '跑贏行業'
    elif any(kw in text.lower() for kw in ['neutral', '中性', 'hold']):
        rating = '中性'
    
    # 提取目標價
    target_price = None
    patterns = [
        r'[Tt]arget [Pp]rice.*?([\d,]+\.?\d*)',
        r'目標價.*?([\d,]+\.?\d*)',
        r'TP.*?([\d,]+\.?\d*)'
    ]
    for pattern in patterns:
        matches = re.findall(pattern, text)
        if matches:
            try:
                target_price = float(matches[0].replace(',', ''))
                if target_price > 100:  # 合理範圍檢查
                    break
            except:
                continue
    
    return broker_name, rating, target_price

def extract_key_points(text):
    """從 PDF 文本提取關鍵要點"""
    
    # 尋找關鍵詞附近嘅內容
    key_point_keywords = [
        r'(?:Key Points|KEY POINTS|Highlights|HIGHLIGHTS|核心觀點|主要觀點|投資要點)[:\s]*([\s\S]{100,500}?)(?=\n\n|Risk|風險|Conclusion)',
        r'(?:Investment Thesis|INVESTMENT THESIS|投資邏輯)[:\s]*([\s\S]{100,500}?)(?=\n\n|Risk|風險)',
        r'(?:Summary|SUMMARY|摘要)[:\s]*([\s\S]{100,500}?)(?=\n\n|Risk|風險)'
    ]
    
    for pattern in key_point_keywords:
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            content = match.group(1).strip()
            # 清理多餘空白
            content = re.sub(r'\s+', ' ', content)
            if len(content) > 50:  # 確保有足夠內容
                return content[:500]  # 限制長度
    
    # 如果冇找到，返回 PDF 前 300 字符作為備用
    clean_text = re.sub(r'\s+', ' ', text[:300]).strip()
    return clean_text if clean_text else "暫無關鍵要點"

def extract_risks(text):
    """從 PDF 文本提取風險因素"""
    
    # 尋找風險相關內容
    risk_keywords = [
        r'(?:Risks|RISKS|Risk Factors|RISK FACTORS|風險|風險因素|投資風險)[:\s]*([\s\S]{100,500}?)(?=\n\n|Conclusion|結論|Appendix)',
        r'(?:Downside Risks|DOWNSIDE RISKS|下行風險)[:\s]*([\s\S]{100,500}?)(?=\n\n|Conclusion)',
        r'(?:Key Risks|KEY RISKS|主要風險)[:\s]*([\s\S]{100,500}?)(?=\n\n|Conclusion)'
    ]
    
    for pattern in risk_keywords:
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            content = match.group(1).strip()
            # 清理多餘空白
            content = re.sub(r'\s+', ' ', content)
            if len(content) > 50:  # 確保有足夠內容
                return content[:500]  # 限制長度
    
    # 如果冇找到，返回默認值
    return "請參考完整報告以了解詳細風險因素"

def extract_release_date(text):
    """從PDF文本中提取發布日期"""
    
    # 常見日期格式模式
    date_patterns = [
        # 英文格式: Date: April 9, 2026 或 9 April 2026
        r'(?:Date|DATE|Published|published|Report Date)[:\s]*(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{4})',
        r'(?:Date|DATE|Published|published|Report Date)[:\s]*((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{1,2},?\s+\d{4})',
        
        # 中文格式: 日期：2026年4月9日 或 2026-04-09
        r'(?:日期|發佈日期|報告日期|发布日期)[:：\s]*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2}[日]?)',
        r'(?:日期|發佈日期|報告日期|发布日期)[:：\s]*(\d{4}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日)',
        
        # ISO格式: 2026-04-09
        r'(\d{4}-\d{2}-\d{2})',
        
        # 其他格式: 09/04/2026 或 04/09/2026
        r'(\d{1,2}/\d{1,2}/\d{4})',
    ]
    
    for pattern in date_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            date_str = match.group(1).strip()
            logger.debug(f"Found date string: {date_str}")
            
            # 嘗試解析各種格式
            try:
                # 嘗試 ISO 格式
                if '-' in date_str and len(date_str) == 10:
                    dt = datetime.strptime(date_str, '%Y-%m-%d')
                    return dt.strftime('%Y-%m-%d')
                
                # 嘗試中文格式
                if '年' in date_str or '月' in date_str:
                    date_str_clean = date_str.replace('年', '-').replace('月', '-').replace('日', '')
                    dt = datetime.strptime(date_str_clean, '%Y-%m-%d')
                    return dt.strftime('%Y-%m-%d')
                
                # 嘗試英文月份格式
                months = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
                         'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}
                
                # April 9, 2026 或 9 April 2026
                parts = date_str.lower().replace(',', '').split()
                if len(parts) >= 3:
                    for i, part in enumerate(parts):
                        if part[:3] in months:
                            month = months[part[:3]]
                            if i > 0 and parts[i-1].isdigit():  # 9 April 2026
                                day = int(parts[i-1])
                                year = int(parts[i+1]) if i+1 < len(parts) and parts[i+1].isdigit() else datetime.now().year
                            elif i < len(parts)-1 and parts[i+1].isdigit():  # April 9, 2026
                                day = int(parts[i+1])
                                year = int(parts[i+2]) if i+2 < len(parts) and parts[i+2].isdigit() else datetime.now().year
                            else:
                                continue
                            
                            dt = datetime(year, month, day)
                            return dt.strftime('%Y-%m-%d')
                
                # DD/MM/YYYY 或 MM/DD/YYYY
                if '/' in date_str:
                    parts = date_str.split('/')
                    if len(parts) == 3:
                        # 假設是 DD/MM/YYYY
                        day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
                        if month <= 12:
                            dt = datetime(year, month, day)
                            return dt.strftime('%Y-%m-%d')
            except Exception as e:
                logger.debug(f"Date parse failed: {e}")
                continue
    
    logger.debug("No valid date found")
    return '-'

def call_nvidia_api(prompt, max_tokens=800, temperature=0.3, retries=2):
    """調用 NVIDIA NIM API 生成 AI 回應（帶重試機制）"""
    for attempt in range(retries + 1):
        try:
            if not NVIDIA_API_KEY or NVIDIA_API_KEY.strip() == '':
                logger.warning("NVIDIA_API_KEY not configured")
                return None
            
            headers = {
                'Authorization': f'Bearer {NVIDIA_API_KEY}',
                'Content-Type': 'application/json',
                'Accept': 'application/json'
            }
            
            payload = {
                'model': NVIDIA_MODEL,
                'messages': [
                    {'role': 'user', 'content': prompt}
                ],
                'max_tokens': max_tokens,
                'temperature': temperature,
                'top_p': 0.7,
                'stream': False
            }
            
            timeout = 120 if attempt == 0 else 90  # 首次嘗試 120 秒，重試 90 秒
            logger.debug(f"Calling NVIDIA NIM API (attempt {attempt + 1}/{retries + 1}) with model: {NVIDIA_MODEL}")
            response = requests.post(NVIDIA_API_URL, headers=headers, json=payload, timeout=timeout)
            
            if response.status_code == 200:
                result = response.json()
                content = result['choices'][0]['message']['content']
                logger.info(f"NVIDIA API success on attempt {attempt + 1}, response length: {len(content)}")
                return content
            elif response.status_code == 429:  # Rate limit
                logger.warning(f"NVIDIA API rate limited (attempt {attempt + 1}), waiting before retry...")
                import time
                time.sleep(5 * (attempt + 1))  # 指數退避
                continue
            else:
                logger.error(f"NVIDIA API error {response.status_code} (attempt {attempt + 1}): {response.text[:300]}")
                if attempt < retries:
                    import time
                    time.sleep(2)  # 等待後重試
                    continue
                return None
                
        except requests.exceptions.Timeout:
            logger.warning(f"NVIDIA API timeout on attempt {attempt + 1}/{retries + 1}")
            if attempt < retries:
                import time
                time.sleep(3)
                continue
            logger.error("NVIDIA API failed after all retries due to timeout")
            return None
        except Exception as e:
            logger.error(f"NVIDIA API call failed (attempt {attempt + 1}): {type(e).__name__}: {str(e)}")
            if attempt < retries:
                import time
                time.sleep(2)
                continue
            import traceback
            logger.error(traceback.format_exc())
            return None
    
    logger.error("NVIDIA API failed after all retry attempts")
    return None

def generate_ai_summary(broker_name, rating, target_price, text):
    """使用 NVIDIA NIM API 生成真正的 AI 摘要"""
    try:
        prompt = f"""你是一位專業的金融分析師。請根據以下券商研報內容,提供簡潔专业的分析摘要。

券商:{broker_name}
評級:{rating}
目標價:HK${'{:.2f}'.format(target_price) if target_price else '未明確'}

研報內容摘要:
{text[:2000]}

請用繁體中文提供:
1. 核心投資觀點(50字以內)
2. 主要風險提示(30字以內)
3. 建議操作策略(30字以內)

格式要求:簡潔明瞭,使用要點式呈現。"""

        ai_content = call_nvidia_api(prompt, max_tokens=500, temperature=0.7)
        
        if ai_content:
            logger.info("AI summary generated successfully via NVIDIA")
            return ai_content.strip()
        else:
            logger.warning("NVIDIA API returned empty")
            return None
    except Exception as e:
        logger.error(f"AI analysis failed: {str(e)}")
        return None

def ensure_traditional_chinese(data):
    """確保所有字段都使用繁體中文，將常見英文翻譯為繁體中文"""
    if not isinstance(data, dict):
        return data
    
    # 常見英文到繁體中文的映射表
    translation_map = {
        # 投資觀點相關
        'Core Investment View': '核心投資觀點',
        'Investment View': '投資觀點',
        'Key Risks': '主要風險',
        'Risk Factors': '風險因素',
        'Recommendation': '建議',
        'Strategy': '策略',
        'Trading Strategy': '操作策略',
        'Action': '行動建議',
        
        # 評級相關
        'Buy': '買入',
        'Hold': '持有',
        'Sell': '賣出',
        'Overweight': '增持',
        'Underweight': '減持',
        'Neutral': '中性',
        'Outperform': '跑贏大市',
        'Underperform': '跑輸大市',
        
        # 行業相關
        'Technology': '科技',
        'Internet': '互聯網',
        'E-commerce': '電商',
        'Gaming': '遊戲',
        'Finance': '金融',
        'Consumer': '消費',
        'Healthcare': '醫療保健',
        'Energy': '能源',
        'Real Estate': '房地產',
        
        # 指數相關
        'Hang Seng Index': '恆生指數',
        'HSI': '恆生指數',
        'Hang Seng Tech Index': '恆生科技指數',
        'HSTECH': '恆生科技指數',
        
        # 其他常見詞
        'Target Price': '目標價',
        'Current Price': '當前價',
        'Upside': '上行空間',
        'Downside': '下行風險',
        'Revenue': '收入',
        'Profit': '利潤',
        'EPS': '每股收益',
        'P/E': '市盈率',
        'Report Date': '報告日期',
        'Analyst': '分析師',
        'Inferred': '推算',
        'Estimated': '估算',
    }
    
    def translate_text(text):
        """翻譯文本中的英文為繁體中文"""
        if not isinstance(text, str):
            return text
        
        result = text
        for eng, chi in translation_map.items():
            result = result.replace(eng, chi)
        
        # 檢查是否仍有大量英文（簡單啟發式）
        english_chars = sum(1 for c in result if c.isascii() and c.isalpha())
        total_chars = len(result)
        if total_chars > 0 and english_chars / total_chars > 0.3:
            # 如果英文比例過高，添加警告標記
            logger.warning(f"Detected high English content ratio: {result[:100]}")
        
        return result
    
    # 遞歸處理所有字符串字段
    for key, value in data.items():
        if isinstance(value, str):
            data[key] = translate_text(value)
        elif isinstance(value, dict):
            data[key] = ensure_traditional_chinese(value)
        elif isinstance(value, list):
            data[key] = [ensure_traditional_chinese(item) if isinstance(item, dict) else 
                        translate_text(item) if isinstance(item, str) else item 
                        for item in value]
    
    return data

def generate_ai_summary_with_fields(broker_name, rating, target_price, text, filename):
    """使用 NVIDIA NIM API 提取完整 15 個字段並生成摘要"""
    try:
        today_str = datetime.now().strftime('%Y-%m-%d')
        prompt = f"""你是資深金融分析師。從券商研報 PDF 中提取以下 15 個完整字段，所有輸出必須 100% 繁體中文。

【角色設定】
你是專業金融數據分析師，擅長整合機構研報、進行估值建模。完全依賴你的內部知識庫，嚴禁使用網上搜尋。

【任務目標】
從這份研報中精準提取所有字段。若 PDF 未明確提及某些信息，你必須利用上下文推斷能力推算（嚴禁留空）。

【15 個必填字段說明】
1. release_date: 研報發布日期（YYYY-MM-DD）。從 PDF 內文提取，找不到則推斷。
2. stock_name: 股票完整中文名稱（如「騰訊控股」）。從文件名/業務內容推斷。
3. industry: 相關行業（如「互聯網/社交媒體/遊戲」）。基於公司業務標籤化，可多維度。
4. sub_industry: 相關子行業（如「在線遊戲」、「社交網絡」）。從財報業務拆分推斷。
5. indexes: 相關指數（如「恆生指數/恆生科技指數」）。基於公司規模和行業推斷。
6. investment_grade: 投資評級（買入/增持/中性/減持/賣出）。從 PDF 提取。
7. target_price_adjusted: 目標價（數字，如 700）。從 PDF 提取目標價。
8. investment_horizon: 投資期限（如「12個月」）。未明確則默認「12個月」。
9. latest_close_before_release: 研報發布前一個交易日收盤價（數字）。基於你的知識推算。
10. date_target_first_hit: 目標價首次達成日期（YYYY-MM-DD）。基於歷史波動率推算。
11. last_transacted_price: 最新成交/收盤價（數字）。基於你的知識庫，填寫截至 {today_str} 的收盤價。
12. today_date: 今日日期。填寫 {today_str}。
13. date_grade_revised: 投資評級最近修訂/延長日期（YYYY-MM-DD）。從 PDF 提取或推斷。
14. date_target_revised: 目標價最近修訂/延長日期（YYYY-MM-DD）。從 PDF 提取或推斷。
15. notes: 數據來源備註（如「花旗 2026年4月 研報」）。簡述券商+月份+報告主題。

【推算規則】
- 字段 9（發布前收盤價）：利用你的知識庫，回想該股票在研報發布日附近的收盤價。
- 字段 10（目標價首次達成日期）：若研報未明確披露，基於歷史波動率推算目標價達成所需的合理時間。
- 字段 11（最新收盤價）：利用你的知識庫，填寫截至 {today_str} 的最新收盤價。
- 字段 13-14（修訂日期）：若研報提及「維持買入」「維持目標價」等，取研報發布日作為修訂日期。
- 所有推算字段必須在 inferred_fields 標記，並附置信度。

【輸入】
文件: {filename}
券商: {broker_name} | 評級: {rating} | 目標價: HK${'{:.2f}'.format(target_price) if target_price else '未明確'}

研報內容:
{text[:3000]}

【輸出格式 - 純淨 JSON，無 markdown 標記】
{{
  "release_date": "YYYY-MM-DD",
  "stock_name": "完整中文名稱",
  "industry": "行業標籤（可多維度，用/分隔）",
  "sub_industry": "子行業標籤",
  "indexes": "相關指數（用/分隔）",
  "investment_grade": "買入/增持/中性/減持/賣出",
  "target_price_adjusted": 700.0,
  "investment_horizon": "12個月",
  "latest_close_before_release": 550.0,
  "date_target_first_hit": "2026-XX-XX",
  "last_transacted_price": 560.0,
  "today_date": "{today_str}",
  "date_grade_revised": "YYYY-MM-DD 或 null",
  "date_target_revised": "YYYY-MM-DD 或 null",
  "notes": "券商+月份+報告主題描述",
  "inferred_fields": ["推斷字段名稱列表"],
  "confidence_scores": {{"latest_close_before_release": 0.7, "date_target_first_hit": 0.5}},
  "ai_summary": "【核心觀點】簡述投資邏輯\n\n【風險提示】關鍵風險因素\n\n【操作建議】具體建議（300字以內，繁體中文）"
}}

確保：所有 15 個字段已填充、價格字段為數字、JSON 格式正確、無英文殘留。"""

        # 檢查 NVIDIA API Key 是否設置
        if not NVIDIA_API_KEY or NVIDIA_API_KEY.strip() == '':
            logger.error("NVIDIA_API_KEY not configured")
            return None, None
        
        # 調用 NVIDIA API（帶重試機制）
        ai_content = call_nvidia_api(prompt, max_tokens=1000, temperature=0.2)
        
        if not ai_content:
            logger.error("NVIDIA API returned empty after retries")
            return None, None
        
        logger.debug(f"Raw NVIDIA response (first 300 chars): {ai_content[:300]}")
        
        # 清理並解析 JSON
        try:
            # 移除 markdown 代碼塊標記
            ai_content_clean = ai_content.strip()
            ai_content_clean = ai_content_clean.replace('```json', '').replace('```', '').strip()
            
            # 嘗試找到 JSON 對象的起始和結束位置
            json_start = ai_content_clean.find('{')
            json_end = ai_content_clean.rfind('}') + 1
            
            if json_start >= 0 and json_end > json_start:
                ai_content_clean = ai_content_clean[json_start:json_end]
            
            extracted_data = json.loads(ai_content_clean)
            logger.info(f"Fields extracted successfully via NVIDIA (attempt with clean JSON)")
            
            # 驗證必要字段是否存在，若缺失則補充
            extracted_data = _validate_and_fill_fields(extracted_data, filename, broker_name)
            
            # 強制檢查並修正語言：確保所有字段都是繁體中文
            extracted_data = ensure_traditional_chinese(extracted_data)
            
            return extracted_data.get('ai_summary', ''), extracted_data
            
        except json.JSONDecodeError as e:
            logger.error(f"JSON parse failed: {e}. Raw content: {ai_content[:200]}")
            # 嘗試修復常見 JSON 錯誤
            fixed_data = _try_fix_json(ai_content)
            if fixed_data:
                logger.info("Successfully fixed JSON parsing error")
                fixed_data = _validate_and_fill_fields(fixed_data, filename, broker_name)
                fixed_data = ensure_traditional_chinese(fixed_data)
                return fixed_data.get('ai_summary', ''), fixed_data
            
            # 若修復失敗，返回原始內容作為摘要
            logger.warning("JSON fix failed, returning raw content as summary")
            return ai_content[:500], _validate_and_fill_fields({}, filename, broker_name)
            
    except Exception as e:
        logger.error(f"Analysis failed with exception: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return None, None





def _validate_and_fill_fields(data, filename, broker_name):
    """驗證並填充缺失字段（智能推斷）"""
    if not isinstance(data, dict):
        data = {}
    
    # 股票名稱推斷
    if not data.get('stock_name') or data['stock_name'] in ['-', '', None]:
        data['stock_name'] = _infer_stock_name(filename)
        if 'stock_name' not in data.get('inferred_fields', []):
            data.setdefault('inferred_fields', []).append('stock_name')
        data.setdefault('confidence_scores', {})['stock_name'] = 0.8
    
    # 行業分類推斷
    if not data.get('industry') or data['industry'] in ['-', '', None]:
        data['industry'] = _infer_industry(data.get('stock_name', ''), filename)
        if 'industry' not in data.get('inferred_fields', []):
            data.setdefault('inferred_fields', []).append('industry')
        data.setdefault('confidence_scores', {})['industry'] = 0.75
    
    # 子行業推斷
    if not data.get('sub_industry') or data['sub_industry'] in ['-', '', None]:
        data['sub_industry'] = _infer_sub_industry(data.get('industry', ''))
        if 'sub_industry' not in data.get('inferred_fields', []):
            data.setdefault('inferred_fields', []).append('sub_industry')
        data.setdefault('confidence_scores', {})['sub_industry'] = 0.7
    
    # 相關指數推斷
    if not data.get('indexes') or data['indexes'] in ['-', '', None]:
        data['indexes'] = _infer_indexes(data.get('stock_name', ''), data.get('industry', ''))
        if 'indexes' not in data.get('inferred_fields', []):
            data.setdefault('inferred_fields', []).append('indexes')
        data.setdefault('confidence_scores', {})['indexes'] = 0.7
    
    # 投資期限默認值
    if not data.get('investment_horizon') or data['investment_horizon'] in ['-', '', None]:
        data['investment_horizon'] = '12個月'
        if 'investment_horizon' not in data.get('inferred_fields', []):
            data.setdefault('inferred_fields', []).append('investment_horizon')
        data.setdefault('confidence_scores', {})['investment_horizon'] = 0.9
    
    # 發布日期默認值
    if not data.get('release_date') or data['release_date'] in ['-', '', None]:
        data['release_date'] = datetime.now().strftime('%Y-%m-%d')
        if 'release_date' not in data.get('inferred_fields', []):
            data.setdefault('inferred_fields', []).append('release_date')
        data.setdefault('confidence_scores', {})['release_date'] = 0.5
    
    return data


def _infer_stock_name(filename):
    """從文件名推斷股票名稱"""
    filename_upper = filename.upper()
    
    stock_mapping = {
        '700': '騰訊控股',
        '0700': '騰訊控股',
        '9988': '阿里巴巴',
        '9618': '京東集團',
        '1810': '小米集團',
        '3690': '美團',
        '1211': '比亞迪股份',
        'TENCENT': '騰訊控股',
        'ALIBABA': '阿里巴巴',
        'JD': '京東集團',
        'XIAOMI': '小米集團',
        'MEITUAN': '美團',
        'BYD': '比亞迪股份'
    }
    
    for key, name in stock_mapping.items():
        if key in filename_upper:
            return name
    
    return '未知股票'


def _infer_industry(stock_name, filename):
    """推斷行業分類"""
    industry_map = {
        '騰訊控股': '互聯網/社交媒體/遊戲/金融科技',
        '阿里巴巴': '電商/雲端計算/物流',
        '京東集團': '電商/物流/零售',
        '小米集團': '消費電子/智能家居/互聯網',
        '美團': '本地生活服務/外賣配送/到店業務',
        '比亞迪股份': '新能源汽車/電池/太陽能'
    }
    
    for key, industry in industry_map.items():
        if key in stock_name:
            return industry
    
    return '綜合企業'


def _infer_sub_industry(industry):
    """推斷子行業"""
    if '遊戲' in industry:
        return '在線遊戲'
    elif '社交' in industry:
        return '社交網絡'
    elif '電商' in industry:
        return '電子商務'
    elif '雲端' in industry or '雲計算' in industry:
        return '雲端服務'
    elif '新能源' in industry or '汽車' in industry:
        return '電動汽車'
    else:
        return '綜合業務'


def _infer_indexes(stock_name, industry):
    """推斷相關指數"""
    indexes = []
    
    # 大型藍籌股
    if any(name in stock_name for name in ['騰訊', '阿里巴巴', '美團', '小米']):
        indexes.append('恆生指數')
    
    # 科技股
    if any(keyword in industry for keyword in ['互聯網', '科技', '電商', '遊戲']):
        indexes.append('恆生科技指數')
    
    # 中國企業
    if any(name in stock_name for name in ['騰訊', '阿里巴巴', '京東', '美團']):
        indexes.append('恆生中國企業指數')
    
    return '/'.join(indexes) if indexes else '恆生指數'


def _try_fix_json(raw_content):
    """嘗試修復常見的 JSON 解析錯誤"""
    try:
        # 移除尾部逗號
        fixed = re.sub(r',\s*}', '}', raw_content)
        fixed = re.sub(r',\s*]', ']', fixed)
        
        # 嘗試解析
        return json.loads(fixed)
    except:
        return None

# ==================== API路由 ====================
# 註：Login/Register endpoint 已移除 - 公開工具無需登入

@app.route('/broker_3quilm/api/analyze', methods=['POST'])
@app.route('/broker_3quilm/api/upload-pdf', methods=['POST'])
def analyze_pdf():
    """分析PDF文件 - 公開訪問，無需登入"""
    try:
        # 公開工具，使用固定用戶ID
        user_id = 1
        
        # Supabase 模式：不需要創建默認用戶
        
        if 'file' not in request.files:
            logger.error("No file uploaded")
            return jsonify({'error': '沒有文件'}), 400
        
        file = request.files['file']
        prompt = request.form.get('prompt', '')
        
        logger.debug(f"Received file: {file.filename}")
        
        # 保存文件（使用 /tmp，Vercel serverless filesystem 是 read-only）
        filename = file.filename
        tmp_folder = '/tmp/pdf_uploads'
        os.makedirs(tmp_folder, exist_ok=True)
        filepath = os.path.join(tmp_folder, filename)
        file.save(filepath)
        logger.debug(f"File saved: {filepath}")
        
        # 解析PDF
        text = parse_pdf(filepath)
        if not text:
            logger.warning(f"PDF parse failed, skipping: {filename}")
            return jsonify({
                'error': f'PDF解析失敗。文件大小: {os.path.getsize(filepath)} bytes。文件可能已損壞或為空文件。請確認文件路徑正確且文件完整。',
                'skipped': True,
                'file_size': os.path.getsize(filepath)
            }), 200  # 返回200而不是500，讓前端知道這是預期行為
        
        logger.debug(f"PDF parsed successfully, text length: {len(text)}")
        
        # 提取基本信息
        broker_name, rating, target_price = extract_broker_info(text)
        
        # 當前價由 AI 提取或前端提供，不再硬編碼
        current_price = None
        
        upside = None
        
        logger.info(f"Extracted - Broker: {broker_name}, Rating: {rating}, Target: {target_price}")
        
        # ========== 去重策略：保留舊記錄作 archive，直接 INSERT 新記錄 ==========
        # 不再刪除舊 record，每次重新分析都新增一條
        # 前端/統計用 get_active_records() 只取每個 pdf_filename 最新的一條
        logger.info(f"Will INSERT new record for {filename} (old records kept as archive)")
        # ========== 去重檢查結束 ==========
        
        # ========== 填充缺失字段 ==========
        company_name = None
        stock_code = None
        
        # 從 PDF 文本提取關鍵要點同風險
        key_points = extract_key_points(text)
        risks = extract_risks(text)
        
        # 如果目標價為 None，設置為 0
        if target_price is None:
            target_price = 0.0
            upside = None
        
        logger.debug(f"Fields - company: {company_name}, stock: {stock_code}, key_points: {len(key_points) if key_points else 0}, risks: {len(risks) if risks else 0}")
        # ========== 填充缺失字段結束 ==========
        
        # 使用真正的 AI 生成摘要並提取完整字段
        ai_summary, extracted_fields = generate_ai_summary_with_fields(broker_name, rating, target_price, text, filename)
        
        # 如果 AI 失敗，返回錯誤
        if not ai_summary:
            logger.error("AI analysis failed, no fallback")
            return jsonify({'error': 'AI 分析失敗，請稍後重試'}), 500
        
        # 保存到 Supabase（含完整 AI 提取字段）
        supabase_data = {
            'user_id': user_id,
            'pdf_filename': filename,
            'broker_name': broker_name,
            'rating': rating,
            'target_price': target_price,
            'current_price': current_price,
            'upside_potential': upside,
            'ai_summary': ai_summary,
            'prompt_used': prompt,
            'company_name': extracted_fields.get('stock_name') if extracted_fields else company_name,
            'stock_code': stock_code,
            'key_points': key_points,
            'risks': risks,
            'chart_path': None,
            'audio_path': None,
            'is_public': 1,
            # AI 提取的 15 個字段
            'release_date': extracted_fields.get('release_date') if extracted_fields else None,
            'stock_name': extracted_fields.get('stock_name') if extracted_fields else None,
            'industry': extracted_fields.get('industry') if extracted_fields else None,
            'sub_industry': extracted_fields.get('sub_industry') if extracted_fields else None,
            'indexes': extracted_fields.get('indexes') if extracted_fields else None,
            'investment_grade': extracted_fields.get('investment_grade') if extracted_fields else None,
            'target_price_adjusted': extracted_fields.get('target_price_adjusted') if extracted_fields else None,
            'investment_horizon': extracted_fields.get('investment_horizon') if extracted_fields else None,
            'latest_close_before_release': extracted_fields.get('latest_close_before_release') if extracted_fields else None,
            'date_target_first_hit': extracted_fields.get('date_target_first_hit') if extracted_fields else None,
            'last_transacted_price': extracted_fields.get('last_transacted_price') if extracted_fields else None,
            'today_date': extracted_fields.get('today_date') if extracted_fields else None,
            'date_grade_revised': extracted_fields.get('date_grade_revised') if extracted_fields else None,
            'date_target_revised': extracted_fields.get('date_target_revised') if extracted_fields else None,
            'notes': extracted_fields.get('notes') if extracted_fields else None,
            'inferred_fields': json.dumps(extracted_fields.get('inferred_fields', [])) if extracted_fields else '[]',
            'confidence_scores': json.dumps(extracted_fields.get('confidence_scores', {})) if extracted_fields else '{}',
            'created_at': datetime.utcnow().isoformat()
        }
        
        result = supabase_request('POST', 'analysis_results', data=supabase_data)
        analysis_id = result[0]['id'] if result and len(result) > 0 else None
        
        logger.info(f"Analysis completed, ID: {analysis_id}")
        
        # 定義路徑變量（用於返回）
        folder_path = UPLOAD_FOLDER
        pdf_folder = UPLOAD_FOLDER
        
        # 計算數據匯總統計
        summary_stats = {
            'avg_target_price': round(target_price, 2) if target_price else None,
            'rating_distribution': {rating: 1},
            'total_analysts': 1
        }
        
        return jsonify({
            'analysis_id': analysis_id,
            'broker_name': broker_name,
            'rating': rating,
            'target_price': target_price,
            'current_price': current_price,
            'upside_potential': upside,
            'ai_summary': ai_summary,
            'pdf_filename': filename,
            'pdf_path': f'{folder_path}/{filename}',
            'upload_path': os.path.join(pdf_folder, filename).replace('\\', '/'),
            # 完整15個字段 - 使用AI提取的數據
            'release_date': extracted_fields.get('release_date', '-'),
            'stock_name': extracted_fields.get('stock_name', '-'),
            'industry': extracted_fields.get('industry', '-'),
            'sub_industry': extracted_fields.get('sub_industry', '-'),
            'indexes': extracted_fields.get('indexes', '-'),
            'target_hit_date': extracted_fields.get('date_target_first_hit', '-'),
            'rating_revised_date': extracted_fields.get('date_grade_revised', '-'),
            'target_revised_date': extracted_fields.get('date_target_revised', '-'),
            'investment_grade': extracted_fields.get('investment_grade', '-'),
            'target_price_adjusted': extracted_fields.get('target_price_adjusted'),
            'latest_close_before_release': extracted_fields.get('latest_close_before_release'),
            'last_transacted_price': extracted_fields.get('last_transacted_price'),
            'inferred_fields': extracted_fields.get('inferred_fields', []),
            # 數據匯總分析
            'summary_stats': summary_stats
        })
    except Exception as e:
        logger.error(f"Analysis error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'伺服器錯誤: {str(e)}'}), 500

@app.route('/broker_3quilm/api/results', methods=['GET'])
def get_results():
    """獲取分析結果 - 公開訪問"""
    user_id = 1  # 公開工具，固定用戶ID
    
    # 從 Supabase 獲取數據
    results = supabase_request('GET', 'analysis_results', query_params=f'user_id=eq.{user_id}&order=created_at.desc')
    
    # 只保留每個 pdf_filename 最新的一條（active record）
    results = get_active_records(results)
    
    if not results:
        return jsonify([])
    
    return jsonify([{
        'id': r.get('id'),
        'pdf_filename': r.get('pdf_filename'),
        'broker_name': r.get('broker_name'),
        'rating': r.get('rating'),
        'target_price': r.get('target_price'),
        'current_price': r.get('current_price'),
        'upside_potential': r.get('upside_potential'),
        'ai_summary': r.get('ai_summary'),
        'created_at': r.get('created_at'),
        # 15 個 AI 提取字段
        'release_date': r.get('release_date'),
        'stock_name': r.get('stock_name') or r.get('company_name'),
        'industry': r.get('industry'),
        'sub_industry': r.get('sub_industry'),
        'indexes': r.get('indexes'),
        'investment_grade': r.get('investment_grade'),
        'target_price_adjusted': r.get('target_price_adjusted'),
        'investment_horizon': r.get('investment_horizon'),
        'latest_close_before_release': r.get('latest_close_before_release'),
        'date_target_first_hit': r.get('date_target_first_hit'),
        'last_transacted_price': r.get('last_transacted_price'),
        'today_date': r.get('today_date'),
        'date_grade_revised': r.get('date_grade_revised'),
        'date_target_revised': r.get('date_target_revised'),
        'notes': r.get('notes'),
        'inferred_fields': r.get('inferred_fields'),
        'confidence_scores': r.get('confidence_scores'),
        'target_hit_date': r.get('date_target_first_hit') or r.get('target_hit_date'),
        'rating_revised_date': r.get('date_grade_revised') or r.get('rating_revised_date'),
        'target_revised_date': r.get('date_target_revised') or r.get('target_revised_date')
    } for r in results])

@app.route('/broker_3quilm/api/feedback', methods=['POST'])
def submit_feedback():
    """提交反饋 - 公開訪問"""
    user_id = 1  # 公開工具，固定用戶ID
    
    data = request.json
    analysis_id = data.get('analysis_id')
    rating = data.get('rating')
    comment = data.get('comment', '')
    suggestions = data.get('suggestions', '')
    
    feedback_data = {
        'user_id': user_id,
        'analysis_id': analysis_id,
        'rating': rating,
        'comment': comment,
        'suggestions': suggestions,
        'created_at': datetime.utcnow().isoformat()
    }
    
    result = supabase_request('POST', 'feedback', data=feedback_data)
    
    return jsonify({'message': '感謝您的反饋!'})

@app.route('/broker_3quilm/api/prompts', methods=['GET'])
def get_prompts():
    """獲取Prompt模板 - 公開訪問"""
    user_id = 1  # 公開工具，固定用戶ID
    
    # 從 Supabase 獲取數據
    results = supabase_request('GET', 'prompt_templates', 
                              query_params=f'or=(user_id.eq.{user_id},is_default.eq.1)&order=is_default.desc,created_at.desc')
    
    if not results:
        return jsonify([])
    
    return jsonify([{
        'id': r.get('id'),
        'template_name': r.get('template_name'),
        'prompt_text': r.get('prompt_text'),
        'is_default': r.get('is_default')
    } for r in results])

@app.route('/broker_3quilm/api/prompts', methods=['POST'])
def save_prompt():
    """保存Prompt模板 - 公開訪問"""
    user_id = 1  # 公開工具，固定用戶ID
    
    data = request.json
    template_name = data.get('template_name')
    prompt_text = data.get('prompt_text')
    
    prompt_data = {
        'user_id': user_id,
        'template_name': template_name,
        'prompt_text': prompt_text,
        'created_at': datetime.utcnow().isoformat()
    }
    
    result = supabase_request('POST', 'prompt_templates', data=prompt_data)
    prompt_id = result[0]['id'] if result and len(result) > 0 else None
    
    return jsonify({'id': prompt_id, 'message': 'Prompt已保存'})



@app.route('/broker_3quilm/dashboard')
def dashboard():
    return send_from_directory('.', 'web/broker_dashboard_v2.html')

@app.route('/broker_3quilm/universal_pdf_dashboard.html')
def universal_pdf_dashboard():
    return send_from_directory('.', 'web/universal_pdf_dashboard.html')

@app.route('/broker_3quilm/')
def root():
    return app.send_static_file('web/login.html')


@app.route('/broker_3quilm/api/charts', methods=['GET'])
def get_charts():
    """獲取圖表數據 - 公開訪問"""
    try:
        # 從 Supabase 獲取數據
        results = supabase_request('GET', 'analysis_results', 
                                  query_params='user_id=eq.1&order=created_at.desc&limit=100')
        
        # 只保留每個 pdf_filename 最新的一條（active record）
        results = get_active_records(results)
        
        if not results:
            return jsonify({
                'brokers': [],
                'ratings': [],
                'upsides': [],
                'distribution': {'Buy': 0, 'Hold': 0, 'Sell': 0}
            }), 200
        
        brokers = [r.get('broker_name') or 'Unknown' for r in results]
        ratings = [r.get('rating') or 'N/A' for r in results]
        upsides = [r.get('upside_potential') if r.get('upside_potential') else 0 for r in results]
        
        buy_count = ratings.count('買入') + ratings.count('Buy')
        hold_count = ratings.count('持有') + ratings.count('Hold') + ratings.count('中性')
        sell_count = ratings.count('賣出') + ratings.count('Sell')
        
        return jsonify({
            'brokers': brokers,
            'ratings': ratings,
            'upsides': upsides,
            'distribution': {'Buy': buy_count, 'Hold': hold_count, 'Sell': sell_count}
        }), 200
    except Exception as e:
        logger.error(f"Charts endpoint error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/broker_3quilm/api/v1/scan/folder', methods=['POST'])
def scan_folder():
    """掃描文件夾中的PDF文件 - 公開訪問"""
    data = request.json
    folder_path = data.get('folder_path', '')
    
    if not folder_path or not os.path.exists(folder_path):
        return jsonify({'error': '文件夾路徑不存在'}), 400
    
    # 查找所有PDF文件
    pdf_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.pdf')]
    
    if not pdf_files:
        return jsonify({'error': '未找到PDF文件', 'total_files': 0}), 200
    
    # 自動分析每個PDF
    analyzed_count = 0
    
    for pdf_file in pdf_files:
        try:
            filepath = os.path.join(folder_path, pdf_file)
            
            # ========== 去重策略：保留舊記錄作 archive，直接 INSERT 新記錄 ==========
            logger.info(f"Will INSERT new record for {pdf_file} (old records kept as archive)")
            # ========== 去重檢查結束 ==========
            
            text = parse_pdf(filepath)
            if text:
                broker_name, rating, target_price = extract_broker_info(text)
                current_price = None
                upside = None
                
                # ========== 使用 AI 提取完整 15 個字段 ==========
                ai_summary, extracted_fields = generate_ai_summary_with_fields(
                    broker_name, rating, target_price, text, pdf_file
                )
                
                # 如果 AI 失敗，跳過此文件
                if not ai_summary:
                    logger.error(f"AI analysis failed for {pdf_file}, skipping")
                    continue
                
                # 提取 AI 返回的完整字段
                company_name = extracted_fields.get('stock_name', None) if extracted_fields else None
                stock_code = None
                key_points = extract_key_points(text)
                risks = extract_risks(text)
                
                if target_price is None:
                    target_price = 0.0
                    upside = None
                # ========== AI 字段提取結束 ==========
                
                # 保存到 Supabase（含完整 AI 提取字段）
                supabase_data = {
                    'user_id': 1,
                    'pdf_filename': pdf_file,
                    'broker_name': broker_name,
                    'rating': rating,
                    'target_price': target_price,
                    'current_price': current_price,
                    'upside_potential': upside,
                    'ai_summary': ai_summary,
                    'prompt_used': '',
                    'company_name': company_name,
                    'stock_code': stock_code,
                    'key_points': key_points,
                    'risks': risks,
                    'chart_path': None,
                    'audio_path': None,
                    'is_public': 1,
                    # AI 提取的 15 個字段
                    'release_date': extracted_fields.get('release_date') if extracted_fields else None,
                    'stock_name': extracted_fields.get('stock_name') if extracted_fields else None,
                    'industry': extracted_fields.get('industry') if extracted_fields else None,
                    'sub_industry': extracted_fields.get('sub_industry') if extracted_fields else None,
                    'indexes': extracted_fields.get('indexes') if extracted_fields else None,
                    'investment_grade': extracted_fields.get('investment_grade') if extracted_fields else None,
                    'target_price_adjusted': extracted_fields.get('target_price_adjusted') if extracted_fields else None,
                    'investment_horizon': extracted_fields.get('investment_horizon') if extracted_fields else None,
                    'latest_close_before_release': extracted_fields.get('latest_close_before_release') if extracted_fields else None,
                    'date_target_first_hit': extracted_fields.get('date_target_first_hit') if extracted_fields else None,
                    'last_transacted_price': extracted_fields.get('last_transacted_price') if extracted_fields else None,
                    'today_date': extracted_fields.get('today_date') if extracted_fields else None,
                    'date_grade_revised': extracted_fields.get('date_grade_revised') if extracted_fields else None,
                    'date_target_revised': extracted_fields.get('date_target_revised') if extracted_fields else None,
                    'notes': extracted_fields.get('notes') if extracted_fields else None,
                    'inferred_fields': json.dumps(extracted_fields.get('inferred_fields', [])) if extracted_fields else '[]',
                    'confidence_scores': json.dumps(extracted_fields.get('confidence_scores', {})) if extracted_fields else '{}',
                    'created_at': datetime.utcnow().isoformat()
                }
                
                logger.info(f"Saving to Supabase (scan): {pdf_file}")
                result = supabase_request('POST', 'analysis_results', data=supabase_data)
                
                if result and len(result) > 0:
                    logger.info(f"Successfully saved to Supabase (scan), ID: {result[0]['id']}")
                    analyzed_count += 1
                else:
                    logger.warning(f"Failed to save to Supabase (scan) for {pdf_file}. Result: {result}")
        except Exception as e:
            logger.error(f"Failed to analyze {pdf_file}: {e}")
            continue
    
    return jsonify({
        'message': '掃描完成',
        'total_files': len(pdf_files),
        'analyzed_files': analyzed_count,
        'details': f'總共 {len(pdf_files)} 個文件，成功分析 {analyzed_count} 個（舊記錄已替換）'
    }), 200

@app.route('/broker_3quilm/api/list-pdfs', methods=['GET'])
def list_pdfs():
    """列出指定文件夾中的所有PDF文件"""
    try:
        # 獲取自定義路徑，如果沒有則使用默認 reports/
        custom_path = request.args.get('path', 'reports')
        
        # 如果是絕對路徑，直接使用；否則相對於項目根目錄
        if os.path.isabs(custom_path):
            pdf_folder = custom_path
        else:
            pdf_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), custom_path)
        
        if not os.path.exists(pdf_folder):
            return jsonify({'error': f'文件夾不存在: {pdf_folder}', 'files': []}), 404
        
        pdf_files = [f for f in os.listdir(pdf_folder) if f.lower().endswith('.pdf')]
        pdf_files.sort()
        
        return jsonify({
            'files': pdf_files,
            'folder_path': pdf_folder.replace('\\', '/')
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'files': []}), 500

@app.route('/broker_3quilm/api/analyze-existing-pdf', methods=['POST'])
def analyze_existing_pdf():
    """分析已存在的PDF文件（不需要上傳）- 公開訪問"""
    try:
        # 公開工具，直接使用固定用戶ID
        user_id = 1
        
        filename = request.form.get('filename')
        folder_path = request.form.get('folder_path', '').strip()  # 獲取自定義文件夾路徑
        if not folder_path:  # 如果為空，使用默認值
            folder_path = 'reports'
        enable_web_search = request.form.get('enable_web_search', 'false').lower() == 'true'  # 是否啟用網絡搜索
        
        if not filename:
            return jsonify({'error': '缺少文件名'}), 400
        
        # 構建文件路徑，支援自定義文件夾
        if os.path.isabs(folder_path):
            pdf_folder = folder_path
        else:
            pdf_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), folder_path)
        
        filepath = os.path.join(pdf_folder, filename)
        
        logger.debug(f"Analyzing PDF - folder: {folder_path}, file: {filename}, path: {filepath}")
        
        if not os.path.exists(filepath):
            return jsonify({'error': f'文件不存在: {filename} (完整路徑: {filepath})'}), 404
        
        # 解析PDF
        logger.debug(f"Starting PDF parse: {filepath}")
        text = parse_pdf(filepath)
        if not text:
            file_size = os.path.getsize(filepath) if os.path.exists(filepath) else 0
            error_msg = f'PDF解析失敗。文件大小: {file_size} bytes。'
            if file_size < 1000:
                error_msg += ' 文件可能已損壞或為空文件。請確認文件路徑正確且文件完整。'
            return jsonify({'error': error_msg}), 500
        
        # 提取基本信息
        broker_name, rating, target_price = extract_broker_info(text)
        
        # 當前價由 AI 提取或前端提供，不再硬編碼
        current_price = None
        
        upside = None
        release_date = extract_release_date(text)
        
        # 使用 AI 生成摘要（包含完整字段提取）
        ai_summary, extracted_fields = generate_ai_summary_with_fields(broker_name, rating, target_price, text, filename)
        
        if not ai_summary:
            logger.error("AI analysis failed, no fallback")
            return jsonify({'error': 'AI 分析失敗，請稍後重試'}), 500
        
        # 如果AI提取成功，使用AI的結果
        final_release_date = extracted_fields.get('release_date', release_date)
        final_stock_name = extracted_fields.get('stock_name', '-')
        final_industry = extracted_fields.get('industry', '-')
        final_sub_industry = extracted_fields.get('sub_industry', '-')
        final_indexes = extracted_fields.get('indexes', '-')
        final_investment_horizon = extracted_fields.get('investment_horizon', '-')
        
        # 準備 Supabase 數據（含完整 15 個 AI 字段）
        supabase_data = {
            'user_id': user_id,
            'pdf_filename': filename,
            'broker_name': broker_name,
            'rating': rating,
            'target_price': target_price,
            'current_price': current_price,
            'upside_potential': upside,
            'ai_summary': ai_summary,
            'prompt_used': '',
            'company_name': extracted_fields.get('stock_name') if extracted_fields else None,
            'release_date': final_release_date,
            'stock_name': final_stock_name,
            'industry': final_industry,
            'sub_industry': final_sub_industry,
            'indexes': final_indexes,
            'investment_grade': extracted_fields.get('investment_grade') if extracted_fields else None,
            'target_price_adjusted': extracted_fields.get('target_price_adjusted') if extracted_fields else None,
            'investment_horizon': final_investment_horizon,
            'latest_close_before_release': extracted_fields.get('latest_close_before_release') if extracted_fields else None,
            'date_target_first_hit': extracted_fields.get('date_target_first_hit') if extracted_fields else None,
            'last_transacted_price': extracted_fields.get('last_transacted_price') if extracted_fields else None,
            'today_date': extracted_fields.get('today_date') if extracted_fields else None,
            'date_grade_revised': extracted_fields.get('date_grade_revised') if extracted_fields else None,
            'date_target_revised': extracted_fields.get('date_target_revised') if extracted_fields else None,
            'notes': extracted_fields.get('notes') if extracted_fields else None,
            'inferred_fields': json.dumps(extracted_fields.get('inferred_fields', [])) if extracted_fields else '[]',
            'confidence_scores': json.dumps(extracted_fields.get('confidence_scores', {})) if extracted_fields else '{}',
            'created_at': datetime.utcnow().isoformat()
        }
        
        # 保存到 Supabase（同步，Vercel Serverless 不支持 background threads）
        # 策略：先刪除舊記錄，再插入新記錄，確保不留舊數據
        analysis_id = None
        try:
            # ========== 去重策略：保留舊記錄作 archive，直接 INSERT 新記錄 ==========
            logger.info(f"Will INSERT new record for {filename} (old records kept as archive)")

            supabase_data['created_at'] = datetime.utcnow().isoformat()

            # 插入新記錄
            result = supabase_request('POST', 'analysis_results', data=supabase_data)
            if result and len(result) > 0:
                analysis_id = result[0]['id']
                logger.info(f"Successfully inserted new record, ID: {analysis_id}")
        except Exception as save_err:
            logger.error(f"Failed to save to Supabase: {save_err}")
        
        # 計算數據匯總統計
        summary_stats = {
            'avg_target_price': round(target_price, 2) if target_price else None,
            'rating_distribution': {rating: 1},  # 單一報告，評級分佈為1
            'total_analysts': 1
        }
        
        return jsonify({
            'analysis_id': analysis_id,
            'broker_name': broker_name,
            'rating': rating,
            'target_price': target_price,
            'current_price': current_price,
            'upside_potential': upside,
            'ai_summary': ai_summary,
            'pdf_filename': filename,
            'pdf_path': f'{folder_path}/{filename}',
            'upload_path': os.path.join(pdf_folder, filename).replace('\\', '/'),
            # 完整15個字段 - 從 PDF 和 AI 提取
            'release_date': final_release_date,
            'stock_name': final_stock_name,
            'industry': final_industry,
            'sub_industry': final_sub_industry,
            'indexes': final_indexes,
            'target_hit_date': extracted_fields.get('date_target_first_hit', '-'),
            'rating_revised_date': extracted_fields.get('date_grade_revised', '-'),
            'target_revised_date': extracted_fields.get('date_target_revised', '-'),
            'investment_horizon': final_investment_horizon,
            'investment_grade': extracted_fields.get('investment_grade', '-'),
            'target_price_adjusted': extracted_fields.get('target_price_adjusted'),
            'latest_close_before_release': extracted_fields.get('latest_close_before_release'),
            'last_transacted_price': extracted_fields.get('last_transacted_price'),
            'inferred_fields': extracted_fields.get('inferred_fields', []),
            # 數據匯總分析
            'summary_stats': summary_stats
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'伺服器錯誤: {str(e)}'}), 500

@app.route('/broker_3quilm/api/chart-data', methods=['GET'])
def get_chart_data():
    """獲取圖表數據（評級分佈、目標價統計等）- 深化數據洞察版本"""
    try:
        # 從 Supabase 獲取所有分析結果
        all_results = supabase_request('GET', 'analysis_results')
        
        # 只保留每個 pdf_filename 最新的一條（active record）
        all_results = get_active_records(all_results)
        
        if not all_results:
            return jsonify({
                'rating_distribution': [],
                'price_statistics': {
                    'total_reports': 0,
                    'average_price': 0,
                    'min_price': 0,
                    'max_price': 0,
                    'median_price': 0,
                    'average_upside': 0,
                    'bull_count': 0,
                    'bear_count': 0,
                    'neutral_count': 0
                },
                'broker_coverage': [],
                'trend_data': [],
                'market_sentiment': ''
            })
        
        # ========== 1. 評級分佈與多空比例 ==========
        rating_count = {}
        bull_ratings = ['買入', '增持', '強烈買入', 'Outperform', 'Overweight', 'Buy']
        bear_ratings = ['賣出', '減持', 'Underperform', 'Underweight', 'Sell']
        neutral_ratings = ['持有', '中性', 'Neutral', 'Hold']
        
        for r in all_results:
            rating = r.get('rating')
            if rating and rating != '' and rating != '-':
                rating_count[rating] = rating_count.get(rating, 0) + 1
        
        rating_data = sorted(rating_count.items(), key=lambda x: x[1], reverse=True)
        
        # 計算多空比例
        bull_count = sum(rating_count.get(r, 0) for r in bull_ratings)
        bear_count = sum(rating_count.get(r, 0) for r in bear_ratings)
        neutral_count = sum(rating_count.get(r, 0) for r in neutral_ratings)
        total_rated = bull_count + bear_count + neutral_count
        
        # ========== 2. 目標價統計（含中位數與上行空間） ==========
        target_prices = [r.get('target_price') for r in all_results if r.get('target_price') and r.get('target_price') > 0]
        upside_potentials = [r.get('upside_potential') for r in all_results if r.get('upside_potential') is not None]
        
        # 計算中位數
        def calculate_median(data):
            if not data:
                return 0
            sorted_data = sorted(data)
            n = len(sorted_data)
            if n % 2 == 0:
                return (sorted_data[n//2 - 1] + sorted_data[n//2]) / 2
            else:
                return sorted_data[n//2]
        
        price_stats = {
            'total_reports': len(target_prices),
            'average_price': round(sum(target_prices) / len(target_prices), 2) if target_prices else 0,
            'median_price': round(calculate_median(target_prices), 2) if target_prices else 0,
            'min_price': round(min(target_prices), 2) if target_prices else 0,
            'max_price': round(max(target_prices), 2) if target_prices else 0,
            'average_upside': round(sum(upside_potentials) / len(upside_potentials), 2) if upside_potentials else 0,
            'bull_count': bull_count,
            'bear_count': bear_count,
            'neutral_count': neutral_count,
            'total_rated': total_rated
        }
        
        # ========== 3. 券商覆蓋 Top 10（含平均目標價、共識度、最新日期） ==========
        broker_stats = {}
        for r in all_results:
            broker = r.get('broker_name')
            if broker and broker != '' and broker != '-':
                if broker not in broker_stats:
                    broker_stats[broker] = {
                        'broker': broker,
                        'count': 0,
                        'target_prices': [],
                        'ratings': [],
                        'latest_date': ''
                    }
                
                broker_stats[broker]['count'] += 1
                
                target_price = r.get('target_price')
                if target_price and target_price > 0:
                    broker_stats[broker]['target_prices'].append(target_price)
                
                rating = r.get('rating')
                if rating and rating != '' and rating != '-':
                    broker_stats[broker]['ratings'].append(rating)
                
                created_at = r.get('created_at')
                if created_at:
                    try:
                        date_str = created_at.split('T')[0] if 'T' in created_at else created_at[:10]
                        if date_str > broker_stats[broker]['latest_date']:
                            broker_stats[broker]['latest_date'] = date_str
                    except:
                        pass
        
        # 計算每個券商的平均目標價和共識度
        broker_data = []
        for broker, stats in broker_stats.items():
            avg_target = round(sum(stats['target_prices']) / len(stats['target_prices']), 2) if stats['target_prices'] else 0
            
            # 計算共識度（最常見的評級）
            if stats['ratings']:
                from collections import Counter
                rating_counter = Counter(stats['ratings'])
                consensus_rating = rating_counter.most_common(1)[0][0]
                consensus_count = rating_counter.most_common(1)[0][1]
                consensus_ratio = round((consensus_count / len(stats['ratings'])) * 100, 1)
            else:
                consensus_rating = '-'
                consensus_ratio = 0
            
            broker_data.append({
                'broker': broker,
                'count': stats['count'],
                'average_target_price': avg_target,
                'consensus_rating': consensus_rating,
                'consensus_ratio': consensus_ratio,
                'latest_date': stats['latest_date']
            })
        
        # 按報告數量排序，取 Top 10
        broker_data = sorted(broker_data, key=lambda x: x['count'], reverse=True)[:10]
        
        # ========== 4. 時間趨勢（最近 30 天） ==========
        from datetime import datetime, timedelta
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        trend_count = {}
        for r in all_results:
            created_at = r.get('created_at')
            if created_at:
                try:
                    date_obj = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                    if date_obj >= thirty_days_ago:
                        date_str = date_obj.strftime('%Y-%m-%d')
                        trend_count[date_str] = trend_count.get(date_str, 0) + 1
                except:
                    pass
        
        trend_data = sorted(trend_count.items())
        
        # ========== 5. 生成 AI 市場情緒摘要 ==========
        sentiment_parts = []
        
        if total_rated > 0:
            # 計算多頭比例
            bull_ratio = round((bull_count / total_rated) * 100, 1)
            
            if bull_ratio >= 80:
                sentiment_level = '極度樂觀'
            elif bull_ratio >= 60:
                sentiment_level = '樂觀'
            elif bull_ratio >= 40:
                sentiment_level = '中性'
            elif bull_ratio >= 20:
                sentiment_level = '審慎'
            else:
                sentiment_level = '悲觀'
            
            sentiment_parts.append(f'目前 {bull_ratio}% 券商給予買入/增持評級')
        
        if price_stats['average_price'] > 0:
            sentiment_parts.append(f'平均目標價 HK${price_stats["average_price"]:.2f}')
        
        if price_stats['average_upside'] > 0:
            sentiment_parts.append(f'較現價有 {price_stats["average_upside"]:.1f}% 上行空間')
        
        if price_stats['min_price'] > 0 and price_stats['max_price'] > 0:
            sentiment_parts.append(f'目標價區間 HK${price_stats["min_price"]:.2f} - HK${price_stats["max_price"]:.2f}')
        
        sentiment_summary = f'市場情緒{sentiment_level}。' + '，'.join(sentiment_parts) + '。' if sentiment_parts else '數據不足以生成市場情緒摘要。'
        
        return jsonify({
            'rating_distribution': [
                {'rating': rating, 'count': count} 
                for rating, count in rating_data
            ],
            'price_statistics': price_stats,
            'broker_coverage': broker_data,
            'trend_data': [
                {'date': date, 'count': count} 
                for date, count in trend_data
            ],
            'market_sentiment': sentiment_summary
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/broker_3quilm/api/archived-reports', methods=['GET'])
def get_archived_reports():
    """獲取所有存檔的舊分析記錄（每個 pdf_filename 的非最新版本）"""
    try:
        all_results = supabase_request('GET', 'analysis_results')
        
        if not all_results:
            return jsonify([])
        
        archived = get_archived_records(all_results)
        
        return jsonify([{
            'id': r.get('id'),
            'pdf_filename': r.get('pdf_filename'),
            'broker_name': r.get('broker_name'),
            'rating': r.get('rating'),
            'target_price': r.get('target_price'),
            'current_price': r.get('current_price'),
            'upside_potential': r.get('upside_potential'),
            'ai_summary': r.get('ai_summary'),
            'key_points': r.get('key_points'),
            'risks': r.get('risks'),
            'stock_name': r.get('stock_name'),
            'created_at': r.get('created_at'),
            'is_archived': True
        } for r in archived])
    except Exception as e:
        logger.error(f"Failed to get archived reports: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/broker_3quilm/api/export-analysis', methods=['GET'])
def export_analysis_report():
    """導出詳細分析報告為 Excel 文件 - 包含完整 15 個字段及智能分表"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        import io
        
        # 從 Supabase 獲取所有數據
        all_results = supabase_request('GET', 'analysis_results')
        
        # 導出時只包含 active record（每個 pdf_filename 最新一條）
        all_results = get_active_records(all_results)
        
        if not all_results:
            return jsonify({'error': '沒有可導出的數據'}), 404
        
        logger.info(f"Exporting {len(all_results)} records to Excel")
        
        # 創建工作簿
        wb = Workbook()
        wb.remove(wb.active)  # 移除默認 sheet
        
        # ========== 樣式定義 ==========
        title_font = Font(name='Microsoft JhengHei', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='667eea', end_color='764ba2', fill_type='solid')
        header_font = Font(name='Microsoft JhengHei', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        normal_font = Font(name='Microsoft JhengHei', size=10)
        warning_font = Font(name='Microsoft JhengHei', size=9, color='FF0000', italic=True)
        warning_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        
        # ========== Sheet 1: 摘要統計 ==========
        ws_summary = wb.create_sheet(title='摘要統計')
        
        # 標題
        ws_summary.merge_cells('A1:B1')
        ws_summary['A1'] = '📊 券商研究報告分析總覽'
        ws_summary['A1'].font = title_font
        ws_summary['A1'].fill = title_fill
        ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_summary.row_dimensions[1].height = 40
        
        # 生成時間
        ws_summary['A2'] = f'生成時間：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws_summary['A2'].font = normal_font
        ws_summary.merge_cells('A2:B2')
        
        # 重要備註
        ws_summary['A3'] = '⚠️ 重要說明：本報告包含所有原始數據，未過濾異常值'
        ws_summary['A3'].font = warning_font
        ws_summary['A3'].fill = warning_fill
        ws_summary.merge_cells('A3:B3')
        ws_summary.row_dimensions[3].height = 25
        
        # 計算基本統計
        target_prices = [r.get('target_price') for r in all_results if r.get('target_price') and r.get('target_price') > 0]
        total_reports = len(target_prices)
        avg_price = sum(target_prices) / len(target_prices) if target_prices else 0
        min_price = min(target_prices) if target_prices else 0
        max_price = max(target_prices) if target_prices else 0
        
        summary_data = [
            ['指標', '數値'],
            ['總報告數', f'{total_reports} 份'],
            ['平均目標價', f'HK${avg_price:.2f}' if avg_price else 'N/A'],
            ['最低目標價', f'HK${min_price:.2f}' if min_price else 'N/A'],
            ['最高目標價', f'HK${max_price:.2f}' if max_price else 'N/A'],
        ]
        
        for row_idx, row_data in enumerate(summary_data, start=5):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 5:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.font = normal_font
                    cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'right', vertical='center')
        
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 20
        
        # ========== Sheet 2: 評級分佈 ==========
        ws_rating = wb.create_sheet(title='評級分佈')
        
        rating_count = {}
        for r in all_results:
            rating = r.get('rating')
            if rating and rating != '' and rating != '-':
                rating_count[rating] = rating_count.get(rating, 0) + 1
        
        rating_data = sorted(rating_count.items(), key=lambda x: x[1], reverse=True)
        total_ratings = sum(count for _, count in rating_data)
        
        ws_rating['A1'] = '評級'
        ws_rating['B1'] = '數量'
        ws_rating['C1'] = '佔比'
        
        for row_idx, (rating, count) in enumerate(rating_data, start=2):
            percentage = (count / total_ratings * 100) if total_ratings > 0 else 0
            ws_rating.cell(row=row_idx, column=1, value=rating).font = normal_font
            ws_rating.cell(row=row_idx, column=2, value=count).font = normal_font
            ws_rating.cell(row=row_idx, column=3, value=f'{percentage:.1f}%').font = normal_font
        
        for col in ['A', 'B', 'C']:
            cell = ws_rating[f'{col}1']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws_rating.column_dimensions[col].width = 15
        
        # ========== Sheet 3: 券商覆蓋排名 ==========
        ws_broker = wb.create_sheet(title='券商覆蓋排名')
        
        broker_count = {}
        for r in all_results:
            broker = r.get('broker_name')
            if broker and broker != '':
                broker_count[broker] = broker_count.get(broker, 0) + 1
        
        broker_data = sorted(broker_count.items(), key=lambda x: x[1], reverse=True)[:20]
        
        ws_broker['A1'] = '排名'
        ws_broker['B1'] = '券商名稱'
        ws_broker['C1'] = '報告數量'
        
        for row_idx, (broker, count) in enumerate(broker_data, start=2):
            ws_broker.cell(row=row_idx, column=1, value=row_idx - 1).font = normal_font
            ws_broker.cell(row=row_idx, column=2, value=broker).font = normal_font
            ws_broker.cell(row=row_idx, column=3, value=count).font = normal_font
        
        for col in ['A', 'B', 'C']:
            cell = ws_broker[f'{col}1']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws_broker.column_dimensions['A'].width = 8
        ws_broker.column_dimensions['B'].width = 25
        ws_broker.column_dimensions['C'].width = 15
        
        # ========== Sheet 4: 完整 15 字段詳細數據 ==========
        ws_full = wb.create_sheet(title='完整數據 (15字段)')
        
        # 定義 15 個字段標題
        full_headers = [
            'Date of Release',           # 1. 發布日期
            'Name of Broker',            # 2. 券商名稱
            'Name of Stock',             # 3. 股票名稱
            'Related Industry',          # 4. 相關行業
            'Related Sub-industry',      # 5. 子行業
            'Related Indexes',           # 6. 相關指數
            'Investment Grade',          # 7. 投資評級
            'Target Price (Adjusted)',   # 8. 目標價
            'Investment Horizon',        # 9. 投資期限
            'Latest Close Before Release', # 10. 發布前收盤價
            'Date of Target First Hit',  # 11. 首次達標日期
            'Last Transacted Price',     # 12. 最新成交價
            "Today's Date",              # 13. 今日日期
            'Date of Grade Revised',     # 14. 評級修訂日期
            'Date of Target Revised'     # 15. 目標價修訂日期
        ]
        
        # 添加額外字段（便於分析）
        extra_headers = [
            'PDF Filename',
            'Current Price',
            'Upside Potential (%)',
            'AI Summary',
            'Analysis Date'
        ]
        
        all_headers = full_headers + extra_headers
        
        # 寫入標題行
        for col_idx, header in enumerate(all_headers, start=1):
            cell = ws_full.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 寫入數據行
        for row_idx, record in enumerate(all_results, start=2):
            # 映射 15 個核心字段
            row_data = [
                record.get('release_date', '-'),                          # 1
                record.get('broker_name', '-'),                           # 2
                record.get('stock_name', '-') or record.get('company_name', '-'),  # 3
                record.get('industry', '-'),                              # 4
                record.get('sub_industry', '-'),                          # 5
                record.get('indexes', '-'),                               # 6
                record.get('rating', '-'),                                # 7
                record.get('target_price', None),                         # 8
                record.get('investment_horizon', '12 months'),            # 9
                record.get('current_price', None),                        # 10
                record.get('target_hit_date', '-'),                       # 11
                record.get('current_price', None),                        # 12 (使用 current_price 作為最新價)
                datetime.now().strftime('%Y-%m-%d'),                      # 13
                record.get('rating_revised_date', '-'),                   # 14
                record.get('target_revised_date', '-')                    # 15
            ]
            
            # 添加額外字段
            extra_data = [
                record.get('pdf_filename', '-'),
                record.get('current_price', None),
                record.get('upside_potential', None),
                record.get('ai_summary', '')[:200] if record.get('ai_summary') else '',  # 限制長度
                record.get('created_at', '-')
            ]
            
            full_row = row_data + extra_data
            
            for col_idx, value in enumerate(full_row, start=1):
                cell = ws_full.cell(row=row_idx, column=col_idx, value=value)
                cell.font = normal_font
                
                # 標註異常價格
                if col_idx == 8 and value and isinstance(value, (int, float)) and value < 100:
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    cell.font = Font(name='Microsoft JhengHei', size=10, color='9C0006', bold=True)
        
        # 設置列寬
        full_widths = [15, 20, 20, 20, 20, 25, 15, 12, 15, 12, 15, 12, 15, 15, 15, 25, 12, 12, 40, 20]
        for col_idx, width in enumerate(full_widths, start=1):
            ws_full.column_dimensions[get_column_letter(col_idx)].width = width
        
        # ========== Sheet 5: 時間趨勢 ==========
        ws_trend = wb.create_sheet(title='時間趨勢')
        
        trend_count = {}
        for r in all_results:
            created_at = r.get('created_at')
            if created_at:
                try:
                    date_obj = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                    date_str = date_obj.strftime('%Y-%m-%d')
                    trend_count[date_str] = trend_count.get(date_str, 0) + 1
                except:
                    pass
        
        trend_data = sorted(trend_count.items())
        
        ws_trend['A1'] = '日期'
        ws_trend['B1'] = '分析數量'
        
        for row_idx, (date, count) in enumerate(trend_data, start=2):
            ws_trend.cell(row=row_idx, column=1, value=date).font = normal_font
            ws_trend.cell(row=row_idx, column=2, value=count).font = normal_font
        
        for col in ['A', 'B']:
            cell = ws_trend[f'{col}1']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws_trend.column_dimensions['A'].width = 15
        ws_trend.column_dimensions['B'].width = 15
        
        # 保存到內存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        filename = f'券商分析報告_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        logger.error(f"Export failed: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# Simple test endpoint (no database required)
# Root endpoint
@app.route('/', methods=['GET'])
def index():
    """根路由 - 返回主頁面"""
    return send_from_directory('.', 'web/universal_pdf_dashboard.html')

@app.route('/api/test', methods=['GET'])
def test_endpoint():
    """簡單測試端點"""
    try:
        ensure_db_initialized()
        return jsonify({
            'status': 'ok',
            'message': 'Backend is running!',
            'timestamp': datetime.now().isoformat(),
            'db_mode': DB_MODE,
            'supabase_configured': bool(SUPABASE_KEY)
        }), 200
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

# Health check endpoint
@app.route('/broker_3quilm/api/health', methods=['GET'])
def health_check():
    """健康檢查端點"""
    try:
        ensure_db_initialized()
        
        # 根據 DB_MODE 選擇不同的檢查方式
        if DB_MODE == 'supabase':
            # Supabase 模式：嘗試查詢一條記錄
            try:
                url = f"{SUPABASE_URL}/rest/v1/analysis_results?select=id&limit=1"
                headers = {
                    'apikey': SUPABASE_KEY,
                    'Authorization': f'Bearer {SUPABASE_KEY}',
                    'Content-Type': 'application/json'
                }
                response = requests.get(url, headers=headers, timeout=5)
                if response.status_code == 200:
                    data = response.json()
                    record_count = len(data) if isinstance(data, list) else 0
                    db_status = 'connected'
                else:
                    record_count = 0
                    db_status = 'error'
            except Exception as e:
                record_count = 0
                db_status = f'error: {str(e)}'
        elif DB_MODE == 'sqlite':
            # SQLite 模式：直接查詢本地數據庫
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM analysis_results')
            record_count = cursor.fetchone()[0]
            conn.close()
            db_status = 'connected'
        else:
            # Memory 模式
            record_count = 0
            db_status = 'memory_mode'
        
        return jsonify({
            'status': 'healthy',
            'timestamp': datetime.now().isoformat(),
            'database': db_status,
            'db_mode': DB_MODE,
            'total_records': record_count,
            'version': '1.0.0'
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

@app.route('/broker_3quilm/api/api-status', methods=['GET'])
def api_status_check():
    """API 狀態檢測 - 顯示連接狀態、響應速度和當前使用的 API"""
    try:
        import time
        
        # 1. 檢查 NVIDIA API
        nvidia_status = {
            'name': 'NVIDIA NIM API',
            'model': NVIDIA_MODEL,
            'configured': bool(NVIDIA_API_KEY and NVIDIA_API_KEY.strip() != ''),
            'status': 'unknown',
            'response_time_ms': None
        }
        
        if nvidia_status['configured']:
            start_time = time.time()
            try:
                test_prompt = "返回 JSON: {\"test\": true}"
                headers = {
                    'Authorization': f'Bearer {NVIDIA_API_KEY}',
                    'Content-Type': 'application/json'
                }
                payload = {
                    'model': NVIDIA_MODEL,
                    'messages': [{'role': 'user', 'content': test_prompt}],
                    'max_tokens': 10,
                    'temperature': 0.1,
                    'stream': False
                }
                
                response = requests.post(NVIDIA_API_URL, headers=headers, json=payload, timeout=30)
                elapsed = (time.time() - start_time) * 1000  # 轉換為毫秒
                
                if response.status_code == 200:
                    nvidia_status['status'] = 'connected'
                    nvidia_status['response_time_ms'] = round(elapsed, 2)
                else:
                    nvidia_status['status'] = 'error'
                    nvidia_status['error'] = f'HTTP {response.status_code}'
                    nvidia_status['response_time_ms'] = round(elapsed, 2)
            except Exception as e:
                elapsed = (time.time() - start_time) * 1000
                nvidia_status['status'] = 'failed'
                nvidia_status['error'] = str(e)[:100]
                nvidia_status['response_time_ms'] = round(elapsed, 2)
        else:
            nvidia_status['status'] = 'not_configured'
        
        # 2. 檢查 OpenRouter API（備用）
        openrouter_status = {
            'name': 'OpenRouter API',
            'configured': bool(OPENROUTER_API_KEY and OPENROUTER_API_KEY.strip() != ''),
            'status': 'unknown',
            'response_time_ms': None
        }
        
        if openrouter_status['configured']:
            start_time = time.time()
            try:
                test_prompt = "返回 JSON: {\"test\": true}"
                headers = {
                    'Authorization': f'Bearer {OPENROUTER_API_KEY}',
                    'Content-Type': 'application/json'
                }
                payload = {
                    'model': 'meta-llama/llama-3.1-8b-instruct',
                    'messages': [{'role': 'user', 'content': test_prompt}],
                    'max_tokens': 10,
                    'temperature': 0.1
                }
                
                response = requests.post(OPENROUTER_API_URL, headers=headers, json=payload, timeout=15)
                elapsed = (time.time() - start_time) * 1000
                
                if response.status_code == 200:
                    openrouter_status['status'] = 'connected'
                    openrouter_status['response_time_ms'] = round(elapsed, 2)
                else:
                    openrouter_status['status'] = 'error'
                    openrouter_status['error'] = f'HTTP {response.status_code}'
                    openrouter_status['response_time_ms'] = round(elapsed, 2)
            except Exception as e:
                elapsed = (time.time() - start_time) * 1000
                openrouter_status['status'] = 'failed'
                openrouter_status['error'] = str(e)[:100]
                openrouter_status['response_time_ms'] = round(elapsed, 2)
        else:
            openrouter_status['status'] = 'not_configured'
        
        # 3. 檢查 Supabase
        supabase_status = {
            'name': 'Supabase Database',
            'url': SUPABASE_URL[:30] + '...' if SUPABASE_URL else None,
            'configured': bool(SUPABASE_URL and SUPABASE_KEY),
            'status': 'unknown',
            'response_time_ms': None
        }
        
        if supabase_status['configured']:
            start_time = time.time()
            try:
                url = f"{SUPABASE_URL}/rest/v1/analysis_results?select=id&limit=1"
                headers = {
                    'apikey': SUPABASE_KEY,
                    'Authorization': f'Bearer {SUPABASE_KEY}'
                }
                response = requests.get(url, headers=headers, timeout=5)
                elapsed = (time.time() - start_time) * 1000
                
                if response.status_code == 200:
                    supabase_status['status'] = 'connected'
                    supabase_status['response_time_ms'] = round(elapsed, 2)
                else:
                    supabase_status['status'] = 'error'
                    supabase_status['error'] = f'HTTP {response.status_code}'
                    supabase_status['response_time_ms'] = round(elapsed, 2)
            except Exception as e:
                elapsed = (time.time() - start_time) * 1000
                supabase_status['status'] = 'failed'
                supabase_status['error'] = str(e)[:100]
                supabase_status['response_time_ms'] = round(elapsed, 2)
        else:
            supabase_status['status'] = 'not_configured'
        
        # 4. 確定主要使用的 API
        primary_api = 'NVIDIA NIM API' if nvidia_status['status'] == 'connected' else \
                     'OpenRouter API' if openrouter_status['status'] == 'connected' else \
                     'None (All APIs failed)'
        
        return jsonify({
            'status': 'success',
            'primary_api': primary_api,
            'apis': {
                'nvidia': nvidia_status,
                'openrouter': openrouter_status,
                'supabase': supabase_status
            },
            'timestamp': datetime.now().isoformat()
        }), 200
        
    except Exception as e:
        logger.error(f"API status check failed: {str(e)}")
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 500

if __name__ == '__main__':
    init_db()
    
    # 僅喺開發環境顯示啟動信息
    if os.environ.get('FLASK_ENV') == 'development':
        logger.info("\n" + "="*60)
        logger.info("🚀 Broker Report Analysis System")
        logger.info("="*60)
        logger.info(f"📍 Server: https://{os.environ.get('VERCEL_URL', 'your-app.vercel.app')}")
        logger.info(f"📍 Health Check: https://{os.environ.get('VERCEL_URL', 'your-app.vercel.app')}/api/health")
        logger.info("="*60 + "\n")
    
    app.run(debug=True, port=62190, use_reloader=False)
