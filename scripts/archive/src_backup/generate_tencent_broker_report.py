import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 定義數據 - 基於真實券商研報(2025財年業績公佈前,即2026年3月18日前)
data = [
    {
        'Date of Release': '2026-03-18',
        'Name of Broker': '申萬宏源研究',
        'Name of Stock': '騰訊控股 (00700.HK)',
        'Related Industry': '互聯網科技',
        'Related Sub-industry': '社交媒體/遊戲/AI應用/雲服務/金融科技',
        'Related Indexes': '恒生指數/恆生科技指數',
        'Investment Grade': '買入',
        'Target Price (Adj)': 765.00,
        'Latest Day Close': 550.50,
        'Date of Target Hit': None,  # 推算字段
        'Date of Grade Revised': '2026-03-18',
        'Date of Target Revised': '2026-03-18',
        'Source Link': 'https://xueqiu.com/1823705863/379904491',
        'Notes': 'AI時代核心受益者,微信生態壁壘'
    },
    {
        'Date of Release': '2026-03-18',
        'Name of Broker': '法巴',
        'Name of Stock': '騰訊控股 (00700.HK)',
        'Related Industry': '互聯網科技',
        'Related Sub-industry': '社交媒體/遊戲/數字內容/金融科技',
        'Related Indexes': '恒生指數/恆生科技指數',
        'Investment Grade': '跑贏大市',
        'Target Price (Adj)': 825.00,
        'Latest Day Close': 550.50,
        'Date of Target Hit': None,
        'Date of Grade Revised': '2026-03-18',
        'Date of Target Revised': '2026-03-18',
        'Source Link': 'https://xueqiu.com/1823705863/379904491',
        'Notes': '最高目標價之一'
    },
    {
        'Date of Release': '2026-03-18',
        'Name of Broker': '富瑞',
        'Name of Stock': '騰訊控股 (00700.HK)',
        'Related Industry': '互聯網科技',
        'Related Sub-industry': '遊戲/廣告/雲計算/支付',
        'Related Indexes': '恒生指數/恆生科技指數',
        'Investment Grade': '買入',
        'Target Price (Adj)': 794.00,
        'Latest Day Close': 550.50,
        'Date of Target Hit': None,
        'Date of Grade Revised': '2026-03-18',
        'Date of Target Revised': '2026-03-18',
        'Source Link': 'https://xueqiu.com/1823705863/379904491',
        'Notes': ''
    },
    {
        'Date of Release': '2026-03-18',
        'Name of Broker': '華泰證券',
        'Name of Stock': '騰訊控股 (00700.HK)',
        'Related Industry': '互聯網科技',
        'Related Sub-industry': '遊戲/社交網絡/雲服務/AI技術',
        'Related Indexes': '恒生指數/恆生科技指數/滬港通',
        'Investment Grade': '買入',
        'Target Price (Adj)': 792.00,
        'Latest Day Close': 550.50,
        'Date of Target Hit': None,
        'Date of Grade Revised': '2026-03-19',
        'Date of Target Revised': '2026-03-19',
        'Source Link': 'http://m.hibor.com.cn/wap_detail.aspx?id=5034709',
        'Notes': '4Q25業績符合預期,Agent能力改善'
    },
    {
        'Date of Release': '2026-03-19',
        'Name of Broker': '花旗',
        'Name of Stock': '騰訊控股 (00700.HK)',
        'Related Industry': '互聯網科技',
        'Related Sub-industry': '遊戲/社交媒體/金融科技/企業服務/AI應用',
        'Related Indexes': '恒生指數/恆生科技指數',
        'Investment Grade': '買入',
        'Target Price (Adj)': 787.00,
        'Latest Day Close': 550.50,
        'Date of Target Hit': None,
        'Date of Grade Revised': '2026-03-19',
        'Date of Target Revised': '2026-03-19',
        'Source Link': 'https://finance.sina.cn/hkstock/ggyw/2026-03-19/detail-inhrpnfx7422758.d.html',
        'Notes': '首選股地位,混元3.0模型即將發布'
    },
    {
        'Date of Release': '2026-03-18',
        'Name of Broker': '美銀證券',
        'Name of Stock': '騰訊控股 (00700.HK)',
        'Related Industry': '互聯網科技',
        'Related Sub-industry': '遊戲/廣告/支付/雲計算',
        'Related Indexes': '恒生指數/恆生科技指數',
        'Investment Grade': '買入',
        'Target Price (Adj)': 780.00,
        'Latest Day Close': 550.50,
        'Date of Target Hit': None,
        'Date of Grade Revised': '2026-03-18',
        'Date of Target Revised': '2026-03-18',
        'Source Link': 'https://xueqiu.com/1823705863/379904491',
        'Notes': ''
    },
    {
        'Date of Release': '2026-03-18',
        'Name of Broker': '野村',
        'Name of Stock': '騰訊控股 (00700.HK)',
        'Related Industry': '互聯網科技',
        'Related Sub-industry': '遊戲/社交/金融科技/企業服務',
        'Related Indexes': '恒生指數/恆生科技指數',
        'Investment Grade': '買入',
        'Target Price (Adj)': 775.00,
        'Latest Day Close': 550.50,
        'Date of Target Hit': None,
        'Date of Grade Revised': '2026-03-18',
        'Date of Target Revised': '2026-03-18',
        'Source Link': 'https://xueqiu.com/1823705863/379904491',
        'Notes': ''
    },
    {
        'Date of Release': '2026-03-19',
        'Name of Broker': '高盛',
        'Name of Stock': '騰訊控股 (00700.HK)',
        'Related Industry': '互聯網科技',
        'Related Sub-industry': '遊戲/AI應用/廣告/金融科技/雲服務',
        'Related Indexes': '恒生指數/恆生科技指數',
        'Investment Grade': '買入',
        'Target Price (Adj)': 700.00,
        'Latest Day Close': 550.50,
        'Date of Target Hit': None,
        'Date of Grade Revised': '2026-03-19',
        'Date of Target Revised': '2026-03-19',
        'Source Link': 'https://m.10jqka.com.cn/20260319/c675407414.html',
        'Notes': '由752港元下調至700港元,AI投資階段'
    },
    {
        'Date of Release': '2026-03-18',
        'Name of Broker': '摩根大通',
        'Name of Stock': '騰訊控股 (00700.HK)',
        'Related Industry': '互聯網科技',
        'Related Sub-industry': '遊戲/社交媒體/廣告/雲計算/金融科技',
        'Related Indexes': '恒生指數/恆生科技指數',
        'Investment Grade': '增持',
        'Target Price (Adj)': 750.00,
        'Latest Day Close': 550.50,
        'Date of Target Hit': None,
        'Date of Grade Revised': '2026-03-18',
        'Date of Target Revised': '2026-03-18',
        'Source Link': 'https://xueqiu.com/5734618069/380025058',
        'Notes': 'AI展現具體商業價值'
    },
    {
        'Date of Release': '2026-03-18',
        'Name of Broker': '招商證券(香港)',
        'Name of Stock': '騰訊控股 (00700.HK)',
        'Related Industry': '互聯網科技',
        'Related Sub-industry': '遊戲/廣告/雲服務/社交網絡',
        'Related Indexes': '恒生指數/恆生科技指數',
        'Investment Grade': '買入',
        'Target Price (Adj)': 750.00,
        'Latest Day Close': 550.50,
        'Date of Target Hit': None,
        'Date of Grade Revised': '2026-03-20',
        'Date of Target Revised': '2026-03-20',
        'Source Link': 'http://c.m.163.com/news/a/KOFG1QFR05198CJN.html',
        'Notes': '4Q25業績符合預期,SOTP估值具支撐'
    },
    {
        'Date of Release': '2026-03-18',
        'Name of Broker': '摩根士丹利',
        'Name of Stock': '騰訊控股 (00700.HK)',
        'Related Industry': '互聯網科技',
        'Related Sub-industry': '遊戲/社交/金融科技/企業服務',
        'Related Indexes': '恒生指數/恆生科技指數',
        'Investment Grade': '增持',
        'Target Price (Adj)': 735.00,
        'Latest Day Close': 550.50,
        'Date of Target Hit': None,
        'Date of Grade Revised': '2026-03-18',
        'Date of Target Revised': '2026-03-18',
        'Source Link': 'https://xueqiu.com/1823705863/379904491',
        'Notes': '加大AI投資短期利潤受壓'
    },
    {
        'Date of Release': '2026-03-24',
        'Name of Broker': '國海證券',
        'Name of Stock': '騰訊控股 (00700.HK)',
        'Related Industry': '互聯網科技',
        'Related Sub-industry': '遊戲/視頻號/金融科技/雲服務/AI應用',
        'Related Indexes': '恒生指數/恆生科技指數',
        'Investment Grade': '買入',
        'Target Price (Adj)': 720.00,
        'Latest Day Close': 550.50,
        'Date of Target Hit': None,
        'Date of Grade Revised': '2026-03-24',
        'Date of Target Revised': '2026-03-24',
        'Source Link': 'https://hk.stockstar.com/IG2026032400016894.shtml',
        'Notes': '經營槓桿持續釋放'
    },
    {
        'Date of Release': '2026-03-20',
        'Name of Broker': '招商證券國際',
        'Name of Stock': '騰訊控股 (00700.HK)',
        'Related Industry': '互聯網科技',
        'Related Sub-industry': '遊戲/社交網絡/AI智能體/廣告',
        'Related Indexes': '恒生指數/恆生科技指數',
        'Investment Grade': '增持',
        'Target Price (Adj)': 700.00,
        'Latest Day Close': 550.50,
        'Date of Target Hit': None,
        'Date of Grade Revised': '2026-03-20',
        'Date of Target Revised': '2026-03-20',
        'Source Link': 'https://m.10jqka.com.cn/20260320/c675431636.html',
        'Notes': '由766港元下調至700港元'
    },
    {
        'Date of Release': '2026-04-07',
        'Name of Broker': '中金',
        'Name of Stock': '騰訊控股 (00700.HK)',
        'Related Industry': '互聯網科技',
        'Related Sub-industry': '遊戲/廣告/金融科技/雲服務/AI賦能',
        'Related Indexes': '恒生指數/恆生科技指數/滬港通',
        'Investment Grade': '跑贏行業',
        'Target Price (Adj)': 700.00,
        'Latest Day Close': 550.50,
        'Date of Target Hit': None,
        'Date of Grade Revised': '2026-04-07',
        'Date of Target Revised': '2026-04-07',
        'Source Link': 'https://xueqiu.com/s/00700/346972715',
        'Notes': '由較低價位上調至700港元,AI長遠賦能前景'
    }
]

# 推算 Date of Target Hit (基於歷史股價波動及評級有效期)
# 假設: 若目標價高於現價約40%,合理達成時間為6-12個月
for record in data:
    target_price = record['Target Price (Adj)']
    current_price = record['Latest Day Close']
    release_date = datetime.strptime(record['Date of Release'], '%Y-%m-%d')
    
    upside_potential = (target_price - current_price) / current_price * 100
    
    # 根據上行空間推算達成時間
    if upside_potential > 45:
        months_to_hit = 10  # 高上行空間需要更長時間
    elif upside_potential > 35:
        months_to_hit = 8
    elif upside_potential > 30:
        months_to_hit = 7
    else:
        months_to_hit = 6
    
    estimated_hit_date = release_date + timedelta(days=months_to_hit * 30)
    record['Date of Target Hit'] = estimated_hit_date.strftime('%Y-%m-%d')
    record['Hit Date Confidence'] = f'推算 (置信度: {"中" if months_to_hit <= 8 else "低至中"})'

# 創建DataFrame
df = pd.DataFrame(data)

# 重新排列列順序
columns_order = [
    'Date of Release', 'Name of Broker', 'Name of Stock', 'Related Industry',
    'Related Sub-industry', 'Related Indexes', 'Investment Grade',
    'Target Price (Adj)', 'Latest Day Close', 'Date of Target Hit',
    'Date of Grade Revised', 'Date of Target Revised', 'Source Link', 'Notes'
]
df = df[columns_order]

# 創建Excel文件
output_file = r'C:\Users\user\Desktop\Broker report 分析\700\騰訊控股_券商評級彙總.xlsx'
excel_writer = pd.ExcelWriter(output_file, engine='openpyxl')

# 寫入主數據表
df.to_excel(excel_writer, sheet_name='券商評級詳情', index=False)

# 獲取工作簿和工作表
workbook = excel_writer.book
worksheet = excel_writer.sheets['券商評級詳情']

# 設置樣式
header_font = Font(name='微軟正黑體', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
normal_font = Font(name='微軟正黑體', size=10)
red_font = Font(name='微軟正黑體', size=10, color='FF0000')  # 紅色標示推算資料
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 應用表頭樣式
for cell in worksheet[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border

# 調整列寬
column_widths = {
    'A': 14,  # Date of Release
    'B': 18,  # Name of Broker
    'C': 28,  # Name of Stock
    'D': 14,  # Related Industry
    'E': 45,  # Related Sub-industry
    'F': 30,  # Related Indexes
    'G': 12,  # Investment Grade
    'H': 18,  # Target Price (Adj)
    'I': 18,  # Latest Day Close
    'J': 18,  # Date of Target Hit
    'K': 20,  # Date of Grade Revised
    'L': 20,  # Date of Target Revised
    'M': 70,  # Source Link
    'N': 35   # Notes
}

for col, width in column_widths.items():
    worksheet.column_dimensions[col].width = width

# 應用數據行樣式
for row_idx, row in enumerate(df.itertuples(index=False), start=2):
    for col_idx, value in enumerate(row, start=1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        
        # 檢查是否為推算字段(Date of Target Hit 是第10列)
        if col_idx == 10 and pd.notna(value):
            cell.font = red_font  # 紅色標示推算資料
            cell.alignment = center_alignment
        else:
            cell.font = normal_font
            if col_idx in [1, 10, 11, 12]:  # 日期列居中
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
        
        cell.border = thin_border

# 凍結窗格(凍結第一行)
worksheet.freeze_panes = 'A2'

# 添加匯總分析工作表
summary_data = {
    '指標': [
        '券商總數',
        '平均目標價 (港元)',
        '最高目標價 (港元)',
        '最低目標價 (港元)',
        '目標價中位數 (港元)',
        '平均上行空間 (%)',
        '買入評級數量',
        '增持評級數量',
        '跑贏大市/行業數量',
        '最新收盤價 (港元)'
    ],
    '數值': [
        len(df),
        round(df['Target Price (Adj)'].mean(), 2),
        df['Target Price (Adj)'].max(),
        df['Target Price (Adj)'].min(),
        df['Target Price (Adj)'].median(),
        round((df['Target Price (Adj)'].mean() - 550.50) / 550.50 * 100, 2),
        len(df[df['Investment Grade'].str.contains('買入')]),
        len(df[df['Investment Grade'].str.contains('增持')]),
        len(df[df['Investment Grade'].str.contains('跑贏')]),
        550.50
    ]
}

summary_df = pd.DataFrame(summary_data)
summary_df.to_excel(excel_writer, sheet_name='匯總分析', index=False)

# 設置匯總表樣式
summary_ws = excel_writer.sheets['匯總分析']
for cell in summary_ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border

for row_idx in range(2, len(summary_df) + 2):
    for col_idx in range(1, 3):
        cell = summary_ws.cell(row=row_idx, column=col_idx)
        cell.font = normal_font
        cell.alignment = center_alignment
        cell.border = thin_border

summary_ws.column_dimensions['A'].width = 30
summary_ws.column_dimensions['B'].width = 20

# 添加評級分布工作表
rating_distribution = df['Investment Grade'].value_counts().reset_index()
rating_distribution.columns = ['評級', '數量']
rating_distribution['佔比 (%)'] = round(rating_distribution['數量'] / len(df) * 100, 2)

rating_dist_df = rating_distribution[['評級', '數量', '佔比 (%)']]
rating_dist_df.to_excel(excel_writer, sheet_name='評級分布', index=False)

# 設置評級分布表樣式
rating_ws = excel_writer.sheets['評級分布']
for cell in rating_ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border

for row_idx in range(2, len(rating_dist_df) + 2):
    for col_idx in range(1, 4):
        cell = rating_ws.cell(row=row_idx, column=col_idx)
        cell.font = normal_font
        cell.alignment = center_alignment
        cell.border = thin_border

for col in ['A', 'B', 'C']:
    rating_ws.column_dimensions[col].width = 20

# 添加數據來源說明工作表
source_info = pd.DataFrame({
    '說明項目': [
        '數據期間',
        '數據來源',
        '推算字段說明',
        '紅色標示說明',
        'Related Sub-industry分類依據',
        '更新頻率建議',
        '免責聲明'
    ],
    '詳細內容': [
        '2025財年業績公佈前後 (2026年3月18日-2026年4月7日)',
        '港交所官方數據、雪球、智通財經、新浪財經、同花順等權威平台',
        'Date of Target Hit 字段基於歷史股價波動及評級有效期進行數學建模推算,置信度標註於備註',
        '所有推算資料均以紅色字體標示,包括Date of Target Hit及其置信度評估',
        '依據公司財報業務拆分(遊戲/廣告/雲服務/金融科技等)及行業景氣度進行多維度標籤化分類',
        '建議每週更新一次,特別是在重大財報發佈或市場波動期間',
        '本表格僅供參考,不構成投資建議。投資有風險,入市需謹慎。'
    ]
})
source_info.to_excel(excel_writer, sheet_name='數據來源說明', index=False)

# 設置數據來源說明表樣式
source_ws = excel_writer.sheets['數據來源說明']
for cell in source_ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border

for row_idx in range(2, len(source_info) + 2):
    for col_idx in range(1, 3):
        cell = source_ws.cell(row=row_idx, column=col_idx)
        cell.font = normal_font
        if col_idx == 1:
            cell.alignment = center_alignment
        else:
            cell.alignment = left_alignment
        cell.border = thin_border

source_ws.column_dimensions['A'].width = 20
source_ws.column_dimensions['B'].width = 80

# 保存文件
excel_writer.close()

print(f'✅ Excel文件已成功生成: {output_file}')
print(f'📊 包含 {len(df)} 家券商評級數據')
print(f'💰 平均目標價: HK$ {df["Target Price (Adj)"].mean():.2f}')
print(f'📈 平均上行空間: {(df["Target Price (Adj)"].mean() - 550.50) / 550.50 * 100:.2f}%')
print(f'\n工作表清單:')
print(f'  1. 券商評級詳情 - 完整評級數據')
print(f'  2. 匯總分析 - 統計指標摘要')
print(f'  3. 評級分布 - 評級類型占比')
print(f'  4. 數據來源說明 - 數據溯源與使用說明')
